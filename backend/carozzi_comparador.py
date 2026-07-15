"""Comparador DNI (cliente) vs C.I. (plataforma) — Carozzi / Molitalia."""

from __future__ import annotations

import csv
import io
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import BinaryIO

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DIGITOS_DNI = 8
COLS_COINCIDENCIAS = ["C.I.", "Nombres", "Apellidos", "Correo"]
COLS_FALTANTES = ["DNI", "Apellidos y Nombres", "Correo cliente"]


def normalizar_dni(valor) -> str:
    """Extrae dígitos y rellena con ceros a la izquierda hasta 8."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    s = str(valor).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digitos = re.sub(r"\D", "", s)
    if not digitos:
        return ""
    if len(digitos) < DIGITOS_DNI:
        digitos = digitos.zfill(DIGITOS_DNI)
    return digitos


def _correo_valido(valor) -> bool:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return False
    s = str(valor).strip()
    if not s:
        return False
    low = s.casefold()
    if low in ("nan", "sin email", "sin correo", "none", "-"):
        return False
    return True


def elegir_correo(correo_carozzi, correo_personal) -> str:
    if _correo_valido(correo_carozzi):
        return str(correo_carozzi).strip()
    if _correo_valido(correo_personal):
        return str(correo_personal).strip()
    return ""


def _detectar_encoding(ruta: Path | str) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(ruta, encoding=encoding) as f:
                f.readline()
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detectar_separador(texto: str) -> str:
    primera = texto.splitlines()[0] if texto else ""
    try:
        dialecto = csv.Sniffer().sniff(texto[:8192], delimiters=";,")
        if dialecto.delimiter in (";", ","):
            return dialecto.delimiter
    except csv.Error:
        pass
    return ";" if primera.count(";") >= primera.count(",") else ","


def _leer_csv(ruta: Path | str) -> pd.DataFrame:
    encoding = _detectar_encoding(ruta)
    with open(ruta, encoding=encoding, errors="replace") as f:
        texto = f.read()
    sep = _detectar_separador(texto)
    kwargs = dict(sep=sep, dtype=str, engine="python", quotechar='"', skipinitialspace=True)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=pd.errors.ParserWarning)
        try:
            return pd.read_csv(io.StringIO(texto), on_bad_lines="warn", **kwargs)
        except TypeError:
            return pd.read_csv(io.StringIO(texto), error_bad_lines=False, warn_bad_lines=False, **kwargs)


def _buscar_columna(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    mapa = {str(c).strip().casefold(): c for c in df.columns}
    for cand in candidatos:
        key = cand.strip().casefold()
        if key in mapa:
            return mapa[key]
    # Coincidencia parcial (ej. "C.I." vs "CI")
    for col in df.columns:
        limpio = re.sub(r"[^a-z0-9]", "", str(col).casefold())
        for cand in candidatos:
            cand_limpo = re.sub(r"[^a-z0-9]", "", cand.casefold())
            if limpio == cand_limpo or cand_limpo in limpio or limpio in cand_limpo:
                return col
    return None


def leer_cliente(ruta: Path | str) -> pd.DataFrame:
    path = Path(ruta)
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(path, dtype=str)
    elif ext == ".csv":
        df = _leer_csv(path)
    else:
        raise ValueError("El archivo del cliente debe ser Excel (.xlsx/.xls) o CSV.")

    col_dni = _buscar_columna(df, ["DNI", "dni", "Documento", "CI", "C.I."])
    if not col_dni:
        raise ValueError("No se encontró la columna DNI en el archivo del cliente.")

    col_nombre = _buscar_columna(df, ["Apellidos y Nombres", "Nombres", "Nombre"])
    col_correo = _buscar_columna(df, ["Correos Molitalia", "Correo", "Email", "Correo cliente"])

    out = pd.DataFrame()
    out["DNI_raw"] = df[col_dni].fillna("").astype(str).str.strip()
    out["DNI_norm"] = out["DNI_raw"].apply(normalizar_dni)
    out["Apellidos y Nombres"] = (
        df[col_nombre].fillna("").astype(str).str.strip() if col_nombre else ""
    )
    out["Correo cliente"] = (
        df[col_correo].fillna("").astype(str).str.strip() if col_correo else ""
    )
    out = out[out["DNI_norm"] != ""].copy()
    return out


def leer_plataforma(ruta: Path | str) -> pd.DataFrame:
    path = Path(ruta)
    if path.suffix.lower() != ".csv":
        raise ValueError("La nómina de plataforma debe ser CSV (.csv).")

    df = _leer_csv(path)
    col_ci = _buscar_columna(df, ["C.I.", "CI", "C.I", "DNI", "Documento"])
    if not col_ci:
        raise ValueError("No se encontró la columna C.I. en la nómina de plataforma.")

    col_nombres = _buscar_columna(df, ["Nombres", "Nombre"])
    col_apellidos = _buscar_columna(df, ["Apellidos", "Apellido"])
    col_carozzi = _buscar_columna(df, ["Correo Carozzi", "CorreoCarozzi"])
    col_personal = _buscar_columna(df, ["Correo personal", "Correo Personal", "Correo Personal"])

    out = pd.DataFrame()
    out["CI_raw"] = df[col_ci].fillna("").astype(str).str.strip()
    out["CI_norm"] = out["CI_raw"].apply(normalizar_dni)
    out["Nombres"] = df[col_nombres].fillna("").astype(str).str.strip() if col_nombres else ""
    out["Apellidos"] = df[col_apellidos].fillna("").astype(str).str.strip() if col_apellidos else ""
    correo_c = df[col_carozzi] if col_carozzi else ""
    correo_p = df[col_personal] if col_personal else ""
    if col_carozzi and col_personal:
        out["Correo"] = [
            elegir_correo(c, p) for c, p in zip(correo_c.fillna(""), correo_p.fillna(""))
        ]
    elif col_carozzi:
        out["Correo"] = [elegir_correo(c, "") for c in correo_c.fillna("")]
    elif col_personal:
        out["Correo"] = [elegir_correo("", p) for p in correo_p.fillna("")]
    else:
        out["Correo"] = ""

    # Descarta filas sin dígitos en C.I. (ej. emails o accesos manuales en esa columna).
    out = out[out["CI_norm"] != ""].copy()
    return out


def _estilo_hoja(ws, fill_header: PatternFill):
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = font_body
            cell.border = border
    ws.row_dimensions[1].height = 26
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def procesar_comparacion(
    ruta_cliente: Path | str,
    ruta_plataforma: Path | str,
    destino: Path | str | BinaryIO,
) -> dict:
    df_cli = leer_cliente(ruta_cliente)
    df_plat = leer_plataforma(ruta_plataforma)

    # Un DNI cliente → primera aparición; plataforma index por C.I. normalizado
    cli_unique = df_cli.drop_duplicates(subset=["DNI_norm"], keep="first")
    plat_by_ci: dict[str, pd.Series] = {}
    for _, row in df_plat.iterrows():
        key = row["CI_norm"]
        if key not in plat_by_ci:
            plat_by_ci[key] = row

    coincidencias = []
    faltantes = []

    for _, row_cli in cli_unique.iterrows():
        dni = row_cli["DNI_norm"]
        row_plat = plat_by_ci.get(dni)
        if row_plat is not None:
            coincidencias.append(
                {
                    "C.I.": dni,
                    "Nombres": row_plat["Nombres"],
                    "Apellidos": row_plat["Apellidos"],
                    "Correo": row_plat["Correo"],
                }
            )
        else:
            faltantes.append(
                {
                    "DNI": dni,
                    "Apellidos y Nombres": row_cli["Apellidos y Nombres"],
                    "Correo cliente": row_cli["Correo cliente"],
                }
            )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_ok = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    fill_miss = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    ws_ok = wb.create_sheet("Coincidencias")
    ws_ok.append(COLS_COINCIDENCIAS)
    for item in coincidencias:
        ws_ok.append([item[c] for c in COLS_COINCIDENCIAS])
    _estilo_hoja(ws_ok, fill_ok)

    ws_miss = wb.create_sheet("Faltantes")
    ws_miss.append(COLS_FALTANTES)
    for item in faltantes:
        ws_miss.append([item[c] for c in COLS_FALTANTES])
    _estilo_hoja(ws_miss, fill_miss)

    wb.save(destino)
    return {
        "coincidencias": len(coincidencias),
        "faltantes": len(faltantes),
        "cliente": len(cli_unique),
        "plataforma": len(plat_by_ci),
    }


def generar_reporte_bytes(
    ruta_cliente: Path | str,
    ruta_plataforma: Path | str,
) -> tuple[bytes, str, dict]:
    buf = io.BytesIO()
    stats = procesar_comparacion(ruta_cliente, ruta_plataforma, buf)
    filename = f"Reporte_Carozzi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf.getvalue(), filename, stats
