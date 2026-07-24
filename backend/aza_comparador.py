"""Comparador de nóminas AZA — conciliación cliente vs plataforma (6 hojas Excel)."""

from __future__ import annotations

import csv
import io
import unicodedata
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import BinaryIO

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CAMPOS = [
    "Rut",
    "Nombres",
    "Apellidos",
    "Correo",
    "Empresa",
    "Nombre de Malla",
    "Malla",
    "Fecha de Ingreso",
    "Nombre del cargo",
    "Fecha Nacimiento",
    "Área Personal",
    "Nacionalidad",
    "División Superior Nombre",
    "División Nombre",
    "Área Superior Nombre",
    "Área Nombre",
    "Centro Costos Nombres",
    "Centro Costo Códigos",
    "Jefe Directo",
    "Rut Jefatura",
]

CAMPOS_EXPORT = [
    "address",
    "firstname",
    "lastname",
    "email",
    "empresa",
    "malla",
    "tipousuario",
    "phone1",
    "city",
    "phone2",
    "yahoo",
    "alternatename",
    "icq",
    "institution",
    "department",
    "description",
    "aim",
    "skype",
    "firstnamephonetic",
    "lastnamephonetic",
    "username",
    "suspended",
]

CAMPOS_EXPORT_ACTUALIZACIONES = CAMPOS_EXPORT[:-1] + ["oldusername", "auth", "suspended"]
CAMPOS_EXPORT_NUEVOS = CAMPOS_EXPORT[:-1] + ["auth", "password", "suspended"]

SEPARADORES_CSV = (";", ",")
MIN_CAMPOS_NOMINA = 15

# Nombres oficiales vigentes (tipousuario = número de malla), según query cliente.
NOMBRE_MALLA_OFICIAL: dict[str, str] = {
    "1": "AZA",
    "2": "AZA Comercial",
    "3": "Filiales",
    "4": "Filiales Comercial",
    "5": "Aza Lideres Operación / Ingenieros / Trainee",
    "6": "EcoAZA",
}

# Reglas de migración: (número legacy, alias de nombre, destino número + nombre oficial).
_REGLAS_MALLAS_LEGACY: list[tuple[str, list[str], tuple[str, str]]] = [
    ("2", ["AZA 2024"], ("1", NOMBRE_MALLA_OFICIAL["1"])),
    ("4", ["AZA Comercial 2024"], ("2", NOMBRE_MALLA_OFICIAL["2"])),
    ("5", ["Filiales"], ("3", NOMBRE_MALLA_OFICIAL["3"])),
    ("6", ["Filiales Comercial"], ("4", NOMBRE_MALLA_OFICIAL["4"])),
    (
        "7",
        [
            "Líderes, Ingenieros y Trainees",
            "Lideres, Ingenieros y Trainees",
            "AZA Líderes Operación/Ingenieros/Trainee",
            "AZA Líderes Operación / Ingenieros / Trainee",
            "AZA Líderes Operación/Ingenieros/Trainees",
            "Líderes Operación/Ingenieros/Trainee",
            "Líderes Operación/Ingenieros/Trainees",
            "Líderes Operación, Ingenieros y Trainees",
            "Líderes Operación, Ingenieros y Trainee",
            "Aza Lideres Operación / Ingenieros / Trainee",
        ],
        ("5", NOMBRE_MALLA_OFICIAL["5"]),
    ),
    ("8", ["EcoAza", "EcoAZA", "ECOAZA"], ("6", NOMBRE_MALLA_OFICIAL["6"])),
]


def normalizar_nombre_malla(nombre) -> str:
    """Comparación insensible a mayúsculas, tildes, separadores y prefijo AZA."""
    s = str(nombre or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.casefold()
    for sep in ("/", ",", ";", "|", "\\", "·", " - ", "-"):
        s = s.replace(sep, " ")
    s = " ".join(s.split())
    if s.startswith("aza "):
        s = s[4:].strip()
    return s


def _construir_mapeo_mallas_legacy() -> dict[tuple[str, str], tuple[str, str]]:
    mapeo: dict[tuple[str, str], tuple[str, str]] = {}
    for numero, alias_list, destino in _REGLAS_MALLAS_LEGACY:
        for alias in alias_list:
            mapeo[(numero, normalizar_nombre_malla(alias))] = destino
    return mapeo


MAPEO_MALLAS_LEGACY = _construir_mapeo_mallas_legacy()

# Mallas eliminadas legacy (sin destino automático). No incluir malla 1 AZA: es la vigente.
_MALLAS_ELIMINADAS_RAW: list[tuple[str, str]] = [
    ("3", "AZA Comercial"),
]


@dataclass(frozen=True)
class ResolucionMalla:
    migrar: bool
    eliminada: bool
    malla_destino: str
    nombre_destino: str


def clave_malla(numero, nombre) -> tuple[str, str]:
    return str(numero or "").strip(), normalizar_nombre_malla(nombre)


MALLAS_ELIMINADAS: frozenset[tuple[str, str]] = frozenset(
    clave_malla(numero, nombre) for numero, nombre in _MALLAS_ELIMINADAS_RAW
)


def _es_malla_7_lideres_legacy(nombre_norm: str) -> bool:
    """Coincidencia flexible para variantes AZA Líderes Operación/Ingenieros/Trainee(s)."""
    return (
        "lideres" in nombre_norm
        and "ingenieros" in nombre_norm
        and "trainee" in nombre_norm
    )


def resolver_malla_plataforma(malla, nombre) -> ResolucionMalla:
    """Detecta malla obsoleta en plataforma y devuelve número/nombre vigentes."""
    key = clave_malla(malla, nombre)
    if key in MAPEO_MALLAS_LEGACY:
        nueva_malla, nuevo_nombre = MAPEO_MALLAS_LEGACY[key]
        return ResolucionMalla(
            migrar=True,
            eliminada=False,
            malla_destino=nueva_malla,
            nombre_destino=nuevo_nombre,
        )
    # Respaldo malla 7: nombre con slashes, prefijo AZA o singular Trainee
    if key[0] == "7" and _es_malla_7_lideres_legacy(key[1]):
        return ResolucionMalla(
            migrar=True,
            eliminada=False,
            malla_destino="5",
            nombre_destino=NOMBRE_MALLA_OFICIAL["5"],
        )
    if key in MALLAS_ELIMINADAS:
        return ResolucionMalla(
            migrar=False,
            eliminada=True,
            malla_destino=str(malla or "").strip(),
            nombre_destino=str(nombre or "").strip(),
        )
    return ResolucionMalla(
        migrar=False,
        eliminada=False,
        malla_destino=str(malla or "").strip(),
        nombre_destino=str(nombre or "").strip(),
    )


def fila_a_exportacion(row, suspended):
    correo = row["Correo"]
    return [
        row["Rut"],
        row["Nombres"],
        row["Apellidos"],
        correo,
        row["Empresa"],
        row["Nombre de Malla"],
        row["Malla"],
        row["Fecha de Ingreso"],
        row["Nombre del cargo"],
        row["Fecha Nacimiento"],
        row["Área Personal"],
        row["Nacionalidad"],
        row["División Superior Nombre"],
        row["División Nombre"],
        row["Área Superior Nombre"],
        row["Área Nombre"],
        row["Centro Costos Nombres"],
        row["Centro Costo Códigos"],
        row["Jefe Directo"],
        row["Rut Jefatura"],
        correo,
        str(suspended),
    ]


def fila_a_exportacion_nuevos(row):
    """Nuevos: auth=saml2 y password igual a address (RUT)."""
    fila = fila_a_exportacion(row, 0)
    address = fila[0]
    return fila[:-1] + ["saml2", address, fila[-1]]


def fila_a_exportacion_actualizacion(row_cli, row_plat):
    fila = fila_a_exportacion(row_cli, 0)
    if valores_diferentes("Correo", row_cli["Correo"], row_plat["Correo"]):
        oldusername = row_plat["Correo"]
        auth = "saml2"
    else:
        oldusername = ""
        auth = ""
    return fila[:-1] + [oldusername, auth, fila[-1]]


def _contar_columnas_linea(linea, separador):
    return len(next(csv.reader([linea], delimiter=separador)))


def _detectar_encoding(ruta: Path | str):
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(ruta, encoding=encoding) as f:
                f.readline()
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _preparar_texto_csv(ruta: Path | str, encoding):
    with open(ruta, encoding=encoding, errors="replace") as f:
        texto = f.read()
    return texto.replace("&amp;", "&").replace("&AMP;", "&")


def _detectar_separador(texto):
    primera_linea = texto.splitlines()[0] if texto else ""
    muestra = texto[:8192]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=";,")
        if dialecto.delimiter in SEPARADORES_CSV:
            return dialecto.delimiter
    except csv.Error:
        pass
    mejor_sep = ";"
    mejor_count = 0
    for sep in SEPARADORES_CSV:
        count = _contar_columnas_linea(primera_linea, sep)
        if count > mejor_count:
            mejor_count = count
            mejor_sep = sep
    return mejor_sep


def _coincidencias_campos_nomina(df):
    if df is None or len(df.columns) <= 1:
        return 0
    nombres_cols = {str(c).strip().lower() for c in df.columns}
    campos_esperados = {c.strip().lower() for c in CAMPOS}
    return len(nombres_cols & campos_esperados)


def _puntuar_dataframe_nomina(df):
    coincidencias = _coincidencias_campos_nomina(df)
    if coincidencias == 0:
        return 0
    return coincidencias * 100 + len(df.columns)


def _leer_csv_desde_texto(texto, sep):
    kwargs = dict(
        sep=sep,
        dtype=str,
        engine="python",
        quotechar='"',
        doublequote=True,
        skipinitialspace=True,
    )
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=pd.errors.ParserWarning)
        try:
            return pd.read_csv(io.StringIO(texto), on_bad_lines="warn", **kwargs)
        except TypeError:
            return pd.read_csv(
                io.StringIO(texto),
                error_bad_lines=False,
                warn_bad_lines=False,
                **kwargs,
            )


def leer_archivo_nomina(ruta: Path | str):
    extension = str(ruta).lower()
    if extension.endswith((".xlsx", ".xls")):
        return pd.read_excel(ruta, dtype=str)

    encoding = _detectar_encoding(ruta)
    texto = _preparar_texto_csv(ruta, encoding)
    sep_detectado = _detectar_separador(texto)
    separadores = []
    for sep in (sep_detectado,) + SEPARADORES_CSV:
        if sep not in separadores:
            separadores.append(sep)

    mejor_df = None
    mejor_puntaje = -1
    ultimo_error = None

    for sep in separadores:
        try:
            df = _leer_csv_desde_texto(texto, sep)
            coincidencias = _coincidencias_campos_nomina(df)
            if coincidencias >= MIN_CAMPOS_NOMINA:
                return df
            puntaje = _puntuar_dataframe_nomina(df)
            if puntaje > mejor_puntaje:
                mejor_puntaje = puntaje
                mejor_df = df
        except Exception as e:
            ultimo_error = e

    if mejor_df is not None and mejor_puntaje >= 300:
        return mejor_df

    nombre = Path(ruta).name
    raise ValueError(
        f"No se pudo leer correctamente el CSV '{nombre}'. Use ; o , como separador "
        f"(Excel) y encabezados como Rut, Nombres, etc. Detalle: {ultimo_error}"
    )


def normalizar_rut(rut):
    if pd.isna(rut) or str(rut).strip() == "":
        return ""
    s = str(rut).strip().upper()
    s = s.replace("\xa0", "").replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace(".", "")
    for guion in ("-", "–", "—", "−"):
        s = s.replace(guion, "")
    return "".join(c for c in s if c.isdigit() or c == "K")


def base_rut(rut_norm):
    if not rut_norm or len(rut_norm) < 2:
        return ""
    return rut_norm[:-1]


def construir_indice_rut_por_base(*dataframes):
    indice = {}
    for df in dataframes:
        for rut in df["Rut"]:
            norm = normalizar_rut(rut)
            if not norm:
                continue
            cuerpo = base_rut(norm)
            if cuerpo and cuerpo not in indice:
                indice[cuerpo] = str(rut).strip()
    return indice


def corregir_rut_jefatura(valor, indice_base_a_rut):
    if pd.isna(valor) or str(valor).strip() == "":
        return valor
    norm = normalizar_rut(valor)
    if not norm:
        return valor
    cuerpo = base_rut(norm)
    if not cuerpo or cuerpo not in indice_base_a_rut:
        return valor
    canonico = indice_base_a_rut[cuerpo]
    if normalizar_rut(canonico) != norm:
        return canonico
    return valor


def aplicar_correccion_ruts_jefatura(df, indice_base_a_rut):
    df = df.copy()
    df["Rut Jefatura"] = df["Rut Jefatura"].apply(
        lambda v: corregir_rut_jefatura(v, indice_base_a_rut)
    )
    return df


def limpiar_dataframe(df):
    mapeo_columnas = {}
    for col in df.columns:
        col_limpia = str(col).strip().lower()
        for campo in CAMPOS:
            if col_limpia == campo.strip().lower():
                mapeo_columnas[col] = campo
                break
    df = df.rename(columns=mapeo_columnas)

    for campo in CAMPOS:
        if campo not in df.columns:
            df[campo] = ""

    df = df[CAMPOS].copy()

    for campo in CAMPOS:
        df[campo] = df[campo].fillna("").astype(str).str.strip()
        df[campo] = df[campo].apply(lambda x: x[:-2] if x.endswith(".0") else x)

    df["Rut_Norm"] = df["Rut"].apply(normalizar_rut)
    df = df[df["Rut_Norm"] != ""]

    return df


def valores_diferentes(campo, val_cli, val_plat):
    if campo in ("Correo", "Rut", "Rut Jefatura"):
        return val_cli.lower() != val_plat.lower()
    return val_cli != val_plat


def construir_fila_destino(row_cli, row_plat, resolucion: ResolucionMalla):
    """Fila objetivo para exportación: migra malla legacy y conserva el resto del cliente."""
    row_destino = row_cli.copy()
    if resolucion.migrar:
        row_destino["Malla"] = resolucion.malla_destino
        row_destino["Nombre de Malla"] = resolucion.nombre_destino
    return row_destino


def analizar_cambios_fila(row_cli, row_plat):
    """
    Analiza una fila de plataforma:
    - Migra mallas legacy (número + nombre) aunque el cliente ya esté actualizado.
    - Compara el resto de campos contra la nómina cliente.
    """
    resolucion = resolver_malla_plataforma(row_plat["Malla"], row_plat["Nombre de Malla"])
    row_destino = construir_fila_destino(row_cli, row_plat, resolucion)

    columnas_con_cambio: list[str] = []

    if resolucion.migrar:
        columnas_con_cambio.extend(["Malla", "Nombre de Malla"])
    elif resolucion.eliminada:
        if valores_diferentes("Malla", row_cli["Malla"], row_plat["Malla"]):
            columnas_con_cambio.append("Malla")
        if valores_diferentes("Nombre de Malla", row_cli["Nombre de Malla"], row_plat["Nombre de Malla"]):
            columnas_con_cambio.append("Nombre de Malla")
    else:
        if valores_diferentes("Malla", row_cli["Malla"], row_plat["Malla"]):
            columnas_con_cambio.append("Malla")
        if valores_diferentes("Nombre de Malla", row_cli["Nombre de Malla"], row_plat["Nombre de Malla"]):
            columnas_con_cambio.append("Nombre de Malla")

    for campo in CAMPOS:
        if campo in ("Nombre de Malla", "Malla"):
            continue
        if valores_diferentes(campo, row_cli[campo], row_plat[campo]):
            columnas_con_cambio.append(campo)

    return columnas_con_cambio, row_destino, resolucion


def analizar_migracion_solo_plataforma(row_plat):
    """Migración de malla para usuarios que solo están en plataforma (sin fila cliente)."""
    resolucion = resolver_malla_plataforma(row_plat["Malla"], row_plat["Nombre de Malla"])
    if not resolucion.migrar:
        return None
    row_destino = row_plat.copy()
    row_destino["Malla"] = resolucion.malla_destino
    row_destino["Nombre de Malla"] = resolucion.nombre_destino
    return row_destino, row_plat


def escribir_hoja_exportacion(ws, filas, suspended, font_header, font_body, fill_header, border_thin):
    ws.append(CAMPOS_EXPORT)
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in filas:
        ws.append(fila_a_exportacion(row, suspended))
        fila_actual = ws.max_row
        for idx in range(1, len(CAMPOS_EXPORT) + 1):
            cell = ws.cell(row=fila_actual, column=idx)
            cell.font = font_body
            cell.border = border_thin


def escribir_hoja_nuevos(ws, filas, font_header, font_body, fill_header, border_thin):
    ws.append(CAMPOS_EXPORT_NUEVOS)
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in filas:
        ws.append(fila_a_exportacion_nuevos(row))
        fila_actual = ws.max_row
        for idx in range(1, len(CAMPOS_EXPORT_NUEVOS) + 1):
            cell = ws.cell(row=fila_actual, column=idx)
            cell.font = font_body
            cell.border = border_thin


def escribir_hoja_actualizaciones(ws, pares_cli_plat, font_header, font_body, fill_header, border_thin):
    ws.append(CAMPOS_EXPORT_ACTUALIZACIONES)
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_cli, row_plat in pares_cli_plat:
        ws.append(fila_a_exportacion_actualizacion(row_cli, row_plat))
        fila_actual = ws.max_row
        for idx in range(1, len(CAMPOS_EXPORT_ACTUALIZACIONES) + 1):
            cell = ws.cell(row=fila_actual, column=idx)
            cell.font = font_body
            cell.border = border_thin


def _aplicar_formato_hojas(wb):
    for ws in wb.worksheets:
        ws.row_dimensions[1].height = 26
        num_cols = ws.max_column
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)
        if ws.max_row >= 1:
            ws.auto_filter.ref = "A1:" + get_column_letter(num_cols) + str(ws.max_row)


def procesar_comparacion(ruta_cliente: Path | str, ruta_plataforma: Path | str, destino: Path | str | BinaryIO):
    df_cli = leer_archivo_nomina(ruta_cliente)
    df_plat = leer_archivo_nomina(ruta_plataforma)

    df_cli = limpiar_dataframe(df_cli)
    df_plat = limpiar_dataframe(df_plat)

    indice_ruts = construir_indice_rut_por_base(df_cli, df_plat)
    df_cli = aplicar_correccion_ruts_jefatura(df_cli, indice_ruts)

    ruts_cli = set(df_cli["Rut_Norm"])
    ruts_plat = set(df_plat["Rut_Norm"])

    ingresos_ruts = ruts_cli - ruts_plat
    salidas_ruts = ruts_plat - ruts_cli
    comunes_ruts = ruts_cli & ruts_plat

    df_ingresos = df_cli[df_cli["Rut_Norm"].isin(ingresos_ruts)].drop(columns=["Rut_Norm"])
    df_salidas = df_plat[df_plat["Rut_Norm"].isin(salidas_ruts)].drop(columns=["Rut_Norm"])

    df_cli_idx = df_cli.set_index("Rut_Norm")
    df_plat_idx = df_plat.set_index("Rut_Norm")

    lista_diferencias = []
    migraciones_solo_plataforma = []

    for rut in comunes_ruts:
        row_cli = df_cli_idx.loc[rut]
        row_plat = df_plat_idx.loc[rut]

        if isinstance(row_cli, pd.DataFrame):
            row_cli = row_cli.iloc[0]
        if isinstance(row_plat, pd.DataFrame):
            row_plat = row_plat.iloc[0]

        columnas_con_cambio, row_destino, _ = analizar_cambios_fila(row_cli, row_plat)

        if columnas_con_cambio:
            lista_diferencias.append((row_destino, row_plat, columnas_con_cambio))

    for rut in salidas_ruts:
        row_plat = df_plat_idx.loc[rut]
        if isinstance(row_plat, pd.DataFrame):
            row_plat = row_plat.iloc[0]
        resultado = analizar_migracion_solo_plataforma(row_plat)
        if resultado:
            migraciones_solo_plataforma.append(resultado)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)

    fill_diff_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_ing_header = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    fill_sal_header = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
    fill_act_header = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    fill_nuevos_header = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    fill_susp_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_cell_modified = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    border_thin = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    ws_diff = wb.create_sheet(title="DIFERENCIAS")
    ws_diff.append(CAMPOS)
    for cell in ws_diff[1]:
        cell.font = font_header
        cell.fill = fill_diff_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_data, _, cols_mod in lista_diferencias:
        valores_fila = [row_data[campo] for campo in CAMPOS]
        ws_diff.append(valores_fila)
        fila_actual = ws_diff.max_row
        for idx, campo in enumerate(CAMPOS, start=1):
            cell = ws_diff.cell(row=fila_actual, column=idx)
            cell.font = font_body
            cell.border = border_thin
            if campo in cols_mod:
                cell.fill = fill_cell_modified

    ws_ing = wb.create_sheet(title="INGRESOS")
    ws_ing.append(CAMPOS)
    for cell in ws_ing[1]:
        cell.font = font_header
        cell.fill = fill_ing_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df_ingresos.iterrows():
        ws_ing.append(list(row))
        fila_actual = ws_ing.max_row
        for idx in range(1, len(CAMPOS) + 1):
            cell = ws_ing.cell(row=fila_actual, column=idx)
            cell.font = font_body
            cell.border = border_thin

    ws_sal = wb.create_sheet(title="SALIDAS")
    ws_sal.append(CAMPOS)
    for cell in ws_sal[1]:
        cell.font = font_header
        cell.fill = fill_sal_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df_salidas.iterrows():
        ws_sal.append(list(row))
        fila_actual = ws_sal.max_row
        for idx in range(1, len(CAMPOS) + 1):
            cell = ws_sal.cell(row=fila_actual, column=idx)
            cell.font = font_body
            cell.border = border_thin

    pares_actualizaciones = [(row_dest, row_plat) for row_dest, row_plat, _ in lista_diferencias]
    pares_actualizaciones.extend(migraciones_solo_plataforma)
    ws_act = wb.create_sheet(title="Actualizaciones")
    escribir_hoja_actualizaciones(
        ws_act, pares_actualizaciones, font_header, font_body, fill_act_header, border_thin
    )

    ws_nuevos = wb.create_sheet(title="Nuevos")
    escribir_hoja_nuevos(
        ws_nuevos,
        [row for _, row in df_ingresos.iterrows()],
        font_header,
        font_body,
        fill_nuevos_header,
        border_thin,
    )

    ws_susp = wb.create_sheet(title="Suspender")
    escribir_hoja_exportacion(
        ws_susp,
        [row for _, row in df_salidas.iterrows()],
        1,
        font_header,
        font_body,
        fill_susp_header,
        border_thin,
    )

    _aplicar_formato_hojas(wb)
    wb.save(destino)

    migraciones_malla = sum(
        1
        for _, row_plat, _ in lista_diferencias
        if resolver_malla_plataforma(row_plat["Malla"], row_plat["Nombre de Malla"]).migrar
    ) + len(migraciones_solo_plataforma)

    return {
        "diferencias": len(lista_diferencias),
        "ingresos": len(df_ingresos),
        "salidas": len(df_salidas),
        "migraciones_malla": migraciones_malla,
    }


def generar_reporte_bytes(ruta_cliente: Path | str, ruta_plataforma: Path | str) -> tuple[bytes, str, dict]:
    buf = io.BytesIO()
    stats = procesar_comparacion(ruta_cliente, ruta_plataforma, buf)
    filename = f"Reporte_AZA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf.getvalue(), filename, stats
