"""Parser y generación CSV para generador Transelec (altas + matriz)."""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

EMAIL_TRANSELEC_REGEX = re.compile(r"[\w.+-]+@transelec\.cl", re.IGNORECASE)
EMAIL_GENERIC_REGEX = re.compile(
    r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
)
PARTICULAS_APELLIDO = {"de", "del", "la", "las", "los", "y", "san", "santa", "von", "van"}

ETIQUETAS_ALTAS = {
    "nombre": "Nombre",
    "rut": "Rut",
    "area": "Area",
    "cargo": "Cargo",
    "fecha inicio contractual": "Fecha inicio contractual",
    "centro de costo- oi": "Centro de Costo- OI",
    "tipo de contrato": "Tipo de Contrato",
    "ubicacion": "Ubicación",
    "ubicación": "Ubicación",
    "requerimientos tecnologicos": "Requerimientos Tecnológicos",
    "requerimientos tecnológicos": "Requerimientos Tecnológicos",
    "elementos de proteccion personal": "Elementos de Protección Personal",
    "elementos de protección personal": "Elementos de Protección Personal",
    "jefatura directa": "Jefatura directa",
    "jornada de trabajo": "Jornada de Trabajo",
}

ETIQUETAS_OBLIGATORIAS = {"nombre", "rut"}


def limpiar_rut(rut_raw: str) -> str:
    return str(rut_raw).replace(".", "").replace(" ", "").strip().upper()


def formatear_palabra_nombre(palabra: str) -> str:
    palabra = palabra.strip()
    if not palabra:
        return palabra
    if len(palabra) <= 3 and palabra.endswith("."):
        return palabra[0].upper() + palabra[1:].lower()
    if "-" in palabra:
        return "-".join(formatear_palabra_nombre(p) for p in palabra.split("-"))
    return palabra[0].upper() + palabra[1:].lower()


def formatear_nombre_partes(partes: list[str]) -> str:
    resultado = []
    for i, parte in enumerate(partes):
        fmt = formatear_palabra_nombre(parte)
        if i > 0 and parte.lower() in PARTICULAS_APELLIDO:
            fmt = parte.lower()
        resultado.append(fmt)
    return " ".join(resultado)


def sugerir_nombre_apellido(nombre_completo: str) -> tuple[str, str]:
    partes = str(nombre_completo).strip().split()
    if not partes:
        return "", ""
    if len(partes) == 1:
        return formatear_nombre_partes(partes), ""
    if len(partes) == 2:
        return formatear_nombre_partes([partes[0]]), formatear_nombre_partes([partes[1]])
    if len(partes) == 3:
        return formatear_nombre_partes([partes[0]]), formatear_nombre_partes(partes[1:])
    nombres = formatear_nombre_partes(partes[:2])
    apellidos = formatear_nombre_partes(partes[2:])
    return nombres, apellidos


def separar_nombre_mayusculas(nombre_completo: str) -> tuple[str, str]:
    nombre_limpio = str(nombre_completo).strip().upper()
    partes = nombre_limpio.split()
    if len(partes) <= 1:
        return nombre_limpio, ""
    if len(partes) == 2:
        return partes[0], partes[1]
    if len(partes) == 3:
        return partes[0], f"{partes[1]} {partes[2]}"
    return f"{partes[0]} {partes[1]}", " ".join(partes[2:])


def _texto_sin_caracteres_rotos(texto: str) -> bool:
    return "\ufffd" not in texto


def leer_lineas_archivo(ruta: Path | str) -> list[str]:
    ruta = Path(ruta)
    if ruta.suffix.lower() == ".csv":
        lineas_respaldo = None
        for encoding in ("utf-8-sig", "cp1252", "latin1", "utf-8"):
            try:
                with open(ruta, encoding=encoding) as archivo:
                    lineas = [linea.rstrip("\n\r") for linea in archivo.readlines()]
                if _texto_sin_caracteres_rotos("".join(lineas)):
                    return lineas
                if lineas_respaldo is None:
                    lineas_respaldo = lineas
            except UnicodeDecodeError:
                continue
        if lineas_respaldo is not None:
            return lineas_respaldo
        raise ValueError("No se pudo leer el CSV con las codificaciones habituales.")

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        sheet = wb.active
        lineas = []
        for row in sheet.iter_rows(values_only=True):
            for valor in row:
                if valor is not None and str(valor).strip():
                    lineas.append(str(valor).strip())
    finally:
        wb.close()
    return lineas


def extraer_email(texto: str) -> str:
    coincidencia = EMAIL_TRANSELEC_REGEX.search(texto)
    return coincidencia.group(0).lower() if coincidencia else ""


def extraer_campo_etiqueta_lineas(lineas: list[str], etiqueta: str) -> str:
    etiqueta_lower = etiqueta.strip().lower()
    for linea in lineas:
        if ";" not in linea:
            continue
        clave, _, valor = linea.partition(";")
        if clave.strip().lower() == etiqueta_lower:
            return valor.strip()
    return ""


def _normalizar_etiqueta_linea(linea: str) -> str | None:
    limpia = linea.strip().rstrip(":").lower()
    if limpia in ETIQUETAS_ALTAS:
        return limpia
    if limpia == "rut":
        return "rut"
    return None


def extraer_campos_desde_lineas(lineas: list[str]) -> dict[str, str]:
    campos: dict[str, str] = {}
    texto_completo = "\n".join(lineas)
    campos["email"] = extraer_email(texto_completo)

    i = 0
    while i < len(lineas):
        linea = lineas[i].strip()
        if not linea:
            i += 1
            continue

        if ";" in linea:
            clave, _, valor = linea.partition(";")
            key = _normalizar_etiqueta_linea(clave.strip())
            if key and valor.strip():
                campos[key] = valor.strip()
            i += 1
            continue

        key = _normalizar_etiqueta_linea(linea)
        if key:
            valores = []
            j = i + 1
            while j < len(lineas):
                siguiente = lineas[j].strip()
                if not siguiente:
                    j += 1
                    continue
                if _normalizar_etiqueta_linea(siguiente) is not None:
                    break
                if EMAIL_TRANSELEC_REGEX.search(siguiente) and key != "email":
                    break
                valores.append(siguiente)
                j += 1
            if valores:
                campos[key] = " ".join(valores)
            i = j
            continue

        i += 1

    if "nombre" not in campos:
        campos["nombre"] = extraer_campo_etiqueta_lineas(lineas, "Nombre")
    if "rut" not in campos:
        campos["rut"] = extraer_campo_etiqueta_lineas(lineas, "Rut")

    return campos


def parsear_solicitud_altas(
    *,
    texto: str | None = None,
    ruta_archivo: Path | str | None = None,
) -> dict[str, Any]:
    if texto:
        lineas = [ln.rstrip("\r") for ln in texto.replace("\r\n", "\n").split("\n")]
    elif ruta_archivo:
        lineas = leer_lineas_archivo(ruta_archivo)
    else:
        raise ValueError("Indica texto o archivo.")

    campos = extraer_campos_desde_lineas(lineas)
    nombre = campos.get("nombre", "").strip()
    rut = campos.get("rut", "").strip()
    email = campos.get("email", "").strip().lower()
    firstname, lastname = sugerir_nombre_apellido(nombre)

    extras = {
        ETIQUETAS_ALTAS[k]: v
        for k, v in campos.items()
        if k not in ("nombre", "rut", "email") and k in ETIQUETAS_ALTAS
    }

    return {
        "email": email,
        "nombre_completo": nombre,
        "rut": rut,
        "firstname": firstname,
        "lastname": lastname,
        "campos_extra": extras,
        "email_es_transelec": bool(email) and email.endswith("@transelec.cl"),
    }


def es_email_valido_matriz(email: str) -> bool:
    if not email or email.lower() == "nan":
        return False
    return bool(EMAIL_GENERIC_REGEX.match(email))


def esta_marcado(valor: Any) -> bool:
    return str(valor).strip().upper() == "X"


def _leer_filas_matriz(ruta: Path | str) -> list[list[str]]:
    ruta = Path(ruta)
    filas: list[list[str]] = []
    if ruta.suffix.lower() == ".csv":
        for encoding in ("utf-8-sig", "cp1252", "latin1", "utf-8"):
            try:
                with open(ruta, encoding=encoding, newline="") as f:
                    reader = csv.reader(f, delimiter=";")
                    for row in reader:
                        filas.append([str(c).strip() if c else "" for c in row])
                return filas
            except UnicodeDecodeError:
                continue
        raise ValueError("No se pudo leer el CSV.")
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            filas.append([str(c).strip() if c is not None else "" for c in row])
    finally:
        wb.close()
    return filas


def procesar_matriz(ruta: Path | str) -> dict[str, Any]:
    filas_raw = _leer_filas_matriz(ruta)
    fecha_hoy = datetime.today().strftime("%d-%m-%Y")
    nombre_grupo = f"Grupo {fecha_hoy}"

    filas_procesadas: list[dict[str, str]] = []
    ruts_vistos: set[str] = set()
    emails_invalidos: list[str] = []
    omitidos_sin_x: list[str] = []
    omitidos_duplicado: list[str] = []

    for row in filas_raw:
        while len(row) < 6:
            row.append("")

        nombre_completo = row[0].strip()
        rut_raw = row[1].strip()
        if rut_raw.lower() in ("nan", "", "rut", "none") or not nombre_completo:
            continue

        email = row[2].strip()
        if email.lower() == "nan":
            email = ""

        col_sub = row[3]
        col_lineas = row[4]
        col_ambas = row[5]
        rut_limpio = limpiar_rut(rut_raw)
        firstname, lastname = separar_nombre_mayusculas(nombre_completo)

        fila: dict[str, str] = {
            "username": rut_limpio,
            "password": rut_limpio,
            "firstname": firstname.upper(),
            "lastname": lastname.upper(),
            "email": email,
            "address": rut_limpio,
            "auth": "manual",
            "institution": "TRANSELEC",
        }

        if esta_marcado(col_ambas):
            fila["course1"] = "Subestaciones"
            fila["group1"] = nombre_grupo
            fila["course2"] = "Líneas de transmisión"
            fila["group2"] = nombre_grupo
        elif esta_marcado(col_lineas):
            fila["course1"] = "Líneas de transmisión"
            fila["group1"] = nombre_grupo
        elif esta_marcado(col_sub):
            fila["course1"] = "Subestaciones"
            fila["group1"] = nombre_grupo
        else:
            omitidos_sin_x.append(f"{rut_limpio} ({nombre_completo})")
            continue

        if rut_limpio in ruts_vistos:
            omitidos_duplicado.append(rut_limpio)
            continue

        if not es_email_valido_matriz(email):
            emails_invalidos.append(f"{rut_limpio} ({nombre_completo}): {email or '(vacío)'}")

        ruts_vistos.add(rut_limpio)
        filas_procesadas.append(fila)

    cols_order = [
        "username", "password", "firstname", "lastname", "email", "address",
        "auth", "institution", "course1", "group1", "course2", "group2",
    ]
    headers = cols_order
    preview_rows = []
    for fila in filas_procesadas[:15]:
        preview_rows.append([fila.get(c, "") for c in cols_order])

    return {
        "nombre_grupo": nombre_grupo,
        "headers": headers,
        "rows": filas_procesadas,
        "preview_rows": preview_rows,
        "total": len(filas_procesadas),
        "warnings": {
            "emails_invalidos": emails_invalidos,
            "omitidos_sin_x": omitidos_sin_x,
            "omitidos_duplicado": omitidos_duplicado,
        },
    }


def construir_fila_alta(datos: dict[str, str], cursos: list[str], grupo: str) -> dict[str, str]:
    rut_limpio = limpiar_rut(datos["rut"])
    email = datos["email"].strip().lower()
    fila = {
        "username": email,
        "password": rut_limpio,
        "address": rut_limpio,
        "firstname": datos["firstname"].strip(),
        "lastname": datos["lastname"].strip(),
        "auth": "saml2",
        "idnumber": rut_limpio,
        "email": email,
        "suspended": "0",
        "institution": "TRANSELEC",
    }
    for i, curso in enumerate(cursos, start=1):
        fila[f"course{i}"] = curso
        fila[f"group{i}"] = grupo
    return fila


def csv_bytes_desde_filas(
    filas: list[dict[str, str]],
    columnas_base: list[str],
    num_cursos: int = 0,
) -> bytes:
    columnas_cursos = []
    for i in range(1, num_cursos + 1):
        columnas_cursos.extend([f"course{i}", f"group{i}"])
    fieldnames = columnas_base + columnas_cursos

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=";", extrasaction="ignore")
    writer.writeheader()
    for fila in filas:
        writer.writerow(fila)
    return buf.getvalue().encode("utf-8-sig")


def generar_csv_alta_bytes(datos: dict[str, str], cursos: list[str], grupo: str) -> tuple[bytes, str]:
    fila = construir_fila_alta(datos, cursos, grupo)
    columnas_base = [
        "username", "password", "address", "firstname", "lastname",
        "auth", "idnumber", "email", "suspended", "institution",
    ]
    content = csv_bytes_desde_filas([fila], columnas_base, len(cursos))
    fecha_str = datetime.today().strftime("%d-%m-%y")
    filename = f"Script_altas - {fecha_str}.csv"
    return content, filename


def generar_csv_matriz_bytes(resultado: dict[str, Any]) -> tuple[bytes, str]:
    filas = resultado["rows"]
    cols_order = resultado["headers"]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=cols_order, delimiter=";", extrasaction="ignore")
    writer.writeheader()
    for fila in filas:
        writer.writerow(fila)
    nombre_grupo = resultado["nombre_grupo"].replace(" ", "_")
    filename = f"Script_{nombre_grupo}.csv"
    return buf.getvalue().encode("utf-8-sig"), filename
