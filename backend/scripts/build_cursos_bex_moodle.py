"""
Genera «cursos bex Moodle.xlsx» unificado:
  - Hoja catálogo (id, shortname, fullname, visible)
  - Una hoja por perfil con solo columna id

Uso: py backend/scripts/build_cursos_bex_moodle.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
CATALOG_XLS = ROOT / "cursos bex Moodle.xls"
LEGACY_MATRIZ = ROOT / "MATRIZ_CURSOS_BEX.xlsx"
OUT_XLSX = ROOT / "cursos bex Moodle.xlsx"

CATALOG_SHEET = "Catalogo Moodle"


def _normalize_id(val) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def load_catalog_from_xls(path: Path) -> list[tuple]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    headers = [str(sh.cell_value(0, c)).strip().lower() for c in range(sh.ncols)]
    id_i = headers.index("id") if "id" in headers else 0
    sn_i = headers.index("shortname") if "shortname" in headers else 2
    fn_i = headers.index("fullname") if "fullname" in headers else 1
    vis_i = headers.index("visible") if "visible" in headers else 3

    rows = []
    for r in range(1, sh.nrows):
        cid = _normalize_id(sh.cell_value(r, id_i))
        if cid is None:
            continue
        shortname = str(sh.cell_value(r, sn_i)).strip()
        fullname = str(sh.cell_value(r, fn_i)).strip() if fn_i < sh.ncols else ""
        visible = sh.cell_value(r, vis_i) if vis_i < sh.ncols else 1
        rows.append((cid, shortname, fullname, visible))
    return rows


def load_profile_ids_from_legacy(path: Path) -> dict[str, list[int]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    profiles: dict[str, list[int]] = {}
    try:
        for sname in wb.sheetnames:
            sheet = wb[sname]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            # Columna ID: encabezado "ID" o segunda columna en formato legacy
            header = [str(c).strip().lower() if c else "" for c in rows[0]]
            id_col = 0
            if "id" in header:
                id_col = header.index("id")
            elif len(header) >= 2:
                id_col = 1

            ids: list[int] = []
            for row in rows[1:]:
                if id_col >= len(row):
                    continue
                cid = _normalize_id(row[id_col])
                if cid is not None:
                    ids.append(cid)
            if ids:
                profiles[sname.strip()] = ids
    finally:
        wb.close()
    return profiles


def build():
    if not CATALOG_XLS.is_file():
        print(f"No existe catálogo: {CATALOG_XLS}")
        sys.exit(1)
    if not LEGACY_MATRIZ.is_file():
        print(f"No existe matriz legacy: {LEGACY_MATRIZ}")
        sys.exit(1)

    catalog = load_catalog_from_xls(CATALOG_XLS)
    profiles = load_profile_ids_from_legacy(LEGACY_MATRIZ)

    wb = Workbook()
    # Catálogo
    ws_cat = wb.active
    ws_cat.title = CATALOG_SHEET[:31]
    ws_cat.append(["id", "shortname", "fullname", "visible"])
    for cid, sn, fn, vis in catalog:
        ws_cat.append([cid, sn, fn, vis])

    # Perfiles (solo id)
    for profile_name, course_ids in profiles.items():
        title = profile_name[:31]
        ws = wb.create_sheet(title=title)
        ws.append(["id"])
        for cid in course_ids:
            ws.append([cid])

    wb.save(OUT_XLSX)
    print(f"Generado: {OUT_XLSX}")
    print(f"  Catálogo: {len(catalog)} cursos")
    print(f"  Perfiles: {len(profiles)} hojas")
    for name, ids in profiles.items():
        print(f"    - {name}: {len(ids)} ids")


if __name__ == "__main__":
    build()
