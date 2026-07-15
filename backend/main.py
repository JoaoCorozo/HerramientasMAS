from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request, Depends, status, Body, Response, BackgroundTasks
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
import json
import os
import csv
import shutil
import tempfile
import zipfile
import unicodedata
from datetime import datetime
import openpyxl
import asyncio
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from database import engine, get_db, Base, SessionLocal
import models
import auth
import config
from deps import get_current_user, require_permission, check_login_rate_limit
from security_utils import (
    safe_upload_filename,
    safe_planilla_filename,
    safe_video_filename,
    read_upload_limited,
    read_video_upload_limited,
    stream_video_upload_to_file,
    escape_html,
    protect_smtp_config,
    expose_smtp_config,
    smtp_config_for_mailer,
    validate_permissions,
    validate_role,
    generic_error_detail,
)

models.Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="Herramientas API",
    docs_url=None if config.IS_PRODUCTION else "/docs",
    redoc_url=None if config.IS_PRODUCTION else "/redoc",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=config.CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logger = logging.getLogger("recordatorios_mailer")

async def dispatcher_background_loop():
    logger.info("Recordatorios mailer background worker started.")
    while True:
        try:
            # Despierta cada 60 segundos
            await asyncio.sleep(60)
            
            db = SessionLocal()
            now = datetime.now()
            
            # Ejecutar despachos si ya son las 9:00 AM o más
            current_date_str = now.strftime("%Y-%m-%d")
            if now.hour >= 9:
                records = db.query(models.AppData).filter(models.AppData.module_name == "recordatorios").all()
                for rec in records:
                    if not rec.payload_json: continue
                    try:
                        tasks_db = json.loads(rec.payload_json)
                    except:
                        continue
                        
                    # Agrupar tareas pendientes por correo de notificación
                    grouped_tasks = {}
                    for date_str, task_list in list(tasks_db.items()):
                        if date_str <= current_date_str:
                            for idx, task in enumerate(task_list):
                                if not task.get("completada", False) and task.get("correo_notificacion") and not task.get("notificado", False):
                                    email = task.get("correo_notificacion").strip().lower()
                                    if email:
                                        if email not in grouped_tasks:
                                            grouped_tasks[email] = []
                                        grouped_tasks[email].append((date_str, idx, task))
                    
                    if not grouped_tasks:
                        continue
                        
                    # Cargar SMTP config de este usuario
                    smtp_rec = db.query(models.AppData).filter(
                        models.AppData.username == rec.username,
                        models.AppData.module_name == "smtp_config"
                    ).first()
                    
                    if not smtp_rec or not smtp_rec.payload_json:
                        continue
                        
                    try:
                        smtp_cfg = smtp_config_for_mailer(json.loads(smtp_rec.payload_json))
                    except:
                        continue
                        
                    if not smtp_cfg or not smtp_cfg.get("host") or not smtp_cfg.get("username") or not smtp_cfg.get("password"):
                        continue
                        
                    dirty = False
                    # Procesar cada destinatario y enviarle su resumen
                    for email, tasks_to_send in grouped_tasks.items():
                        try:
                            tasks_html = ""
                            for date_str, idx, t in tasks_to_send:
                                tasks_html += f"""
                                <div style="margin-bottom: 20px; padding: 20px; border: 1px solid #e4e4e7; border-radius: 8px; background-color: #fafafa; box-shadow: 0 1px 3px rgba(0,0,0,0.05);">
                                    <div style="margin-bottom: 12px; border-bottom: 1px solid #f4f4f5; padding-bottom: 8px; display: flex; justify-content: space-between; align-items: center;">
                                        <span style="font-size: 16px; font-weight: bold; color: #8b5cf6;">📌 {escape_html(t.get('titulo', 'Tarea'))}</span>
                                        <span style="font-size: 12px; color: #a1a1aa; background-color: #f4f4f5; padding: 4px 8px; border-radius: 4px; font-weight: bold;">📅 {escape_html(date_str)}</span>
                                    </div>
                                    <p style="margin: 0 0 12px 0; font-size: 14px; color: #3f3f46; line-height: 1.5;">{escape_html(t.get('cuerpo_mail') or t.get('detalle', 'Sin detalles adicionales.'))}</p>
                                    <table style="width: 100%; border-collapse: collapse; font-size: 13px;">
                                """
                                if t.get("curso"):
                                    tasks_html += f"""
                                        <tr>
                                            <td style="padding: 4px 0; font-weight: bold; color: #71717a; width: 100px;">📚 Curso ID:</td>
                                            <td style="padding: 4px 0; color: #18181b;">{escape_html(t.get('curso'))}</td>
                                        </tr>
                                    """
                                if t.get("grupo"):
                                    tasks_html += f"""
                                        <tr>
                                            <td style="padding: 4px 0; font-weight: bold; color: #71717a;">👥 Grupo:</td>
                                            <td style="padding: 4px 0; color: #18181b;">{escape_html(t.get('grupo'))}</td>
                                        </tr>
                                    """
                                if t.get("asunto"):
                                    tasks_html += f"""
                                        <tr>
                                            <td style="padding: 4px 0; font-weight: bold; color: #71717a;">📝 Asunto:</td>
                                            <td style="padding: 4px 0; color: #18181b;">{escape_html(t.get('asunto'))}</td>
                                        </tr>
                                    """
                                tasks_html += """
                                    </table>
                                </div>
                                """

                            html_content = f"""
                            <html>
                            <body style="font-family: Arial, sans-serif; background-color: #f4f4f5; padding: 20px; margin: 0;">
                                <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; border-radius: 12px; border: 1px solid #e4e4e7; overflow: hidden; box-shadow: 0 4px 10px rgba(0,0,0,0.05);">
                                    <div style="background-color: #8b5cf6; padding: 24px; color: #ffffff; text-align: center;">
                                        <h2 style="margin: 0; font-size: 24px; font-weight: bold; letter-spacing: -0.5px;">🔔 Resumen de Tareas Pendientes</h2>
                                        <p style="margin: 6px 0 0 0; opacity: 0.9; font-size: 14px;">Plataforma de Herramientas BEX</p>
                                    </div>
                                    <div style="padding: 24px; color: #18181b;">
                                        <p style="font-size: 15px; color: #3f3f46; margin-top: 0; margin-bottom: 20px;">Tienes las siguientes tareas programadas para hoy:</p>
                                        
                                        {tasks_html}
                                        
                                        <hr style="border: 0; border-top: 1px solid #e4e4e7; margin: 24px 0;">
                                        <p style="font-size: 12px; color: #a1a1aa; text-align: center; margin-bottom: 0; line-height: 1.5;">
                                            Este es un recordatorio automático consolidado enviado a las 9:00 AM.<br>
                                            Por favor, no respondas a este correo.
                                        </p>
                                    </div>
                                </div>
                            </body>
                            </html>
                            """
                            
                            msg = MIMEMultipart('alternative')
                            msg['Subject'] = f"🔔 Resumen de Tareas Pendientes ({len(tasks_to_send)} Tareas)"
                            sender_name = smtp_cfg.get("sender_name") or "Plataforma Herramientas"
                            sender_addr = smtp_cfg.get("sender_email") or smtp_cfg.get("username")
                            msg['From'] = f"{sender_name} <{sender_addr}>"
                            msg['To'] = email
                            
                            msg.attach(MIMEText(html_content, 'html'))
                            
                            # Conexión SMTP
                            srv = smtplib.SMTP(smtp_cfg["host"], int(smtp_cfg["port"]))
                            if smtp_cfg.get("use_tls", True) or int(smtp_cfg.get("port", 587)) == 587:
                                srv.starttls()
                            srv.login(smtp_cfg["username"], smtp_cfg["password"])
                            srv.sendmail(sender_addr, email, msg.as_string())
                            srv.quit()
                            
                            # Marcar como notificado
                            for date_str, idx, task in tasks_to_send:
                                tasks_db[date_str][idx]["notificado"] = True
                            dirty = True
                            logger.info(f"Consolidated notification email sent to {email} with {len(tasks_to_send)} tasks.")
                        except Exception as send_err:
                            logger.error(f"Error sending consolidated email to {email}: {send_err}")
                            
                    if dirty:
                        rec.payload_json = json.dumps(tasks_db)
                        db.commit()
                        
            db.close()
        except Exception as loop_err:
            logger.error(f"Error in dispatcher background loop: {loop_err}")

DEFAULT_ADMIN_PERMISSIONS = json.dumps([
    "comparador", "rut", "textos", "capacitaciones",
    "enlaces", "recordatorios", "generador", "consulta_cursos",
    "usuarios_duplicados", "compresor_video",
])


@app.on_event("startup")
def startup_event():
    db = next(get_db())
    admin_user = db.query(models.User).filter(models.User.username == "admin").first()
    if not admin_user and config.BOOTSTRAP_ADMIN_PASSWORD:
        admin_user = models.User(
            username="admin",
            hashed_password=auth.get_password_hash(config.BOOTSTRAP_ADMIN_PASSWORD),
            role="superadmin",
            permissions_json=DEFAULT_ADMIN_PERMISSIONS,
        )
        db.add(admin_user)
        db.commit()
        logger.info("Usuario admin creado (BOOTSTRAP_ADMIN_PASSWORD). Cambie la contraseña tras el primer acceso.")
    try:
        from matriz_db import ensure_matriz_seeded

        ensure_matriz_seeded(db)
    except Exception as e:
        logger.warning("No se pudo importar matriz Moodle al iniciar: %s", e)
    try:
        from transelec_db import ensure_transelec_seeded

        ensure_transelec_seeded(db)
    except Exception as e:
        logger.warning("No se pudo inicializar catálogo Transelec: %s", e)
    finally:
        db.close()
    asyncio.create_task(dispatcher_background_loop())


def set_auth_cookie(response: Response, access_token: str) -> None:
    response.set_cookie(
        key=config.COOKIE_NAME,
        value=access_token,
        httponly=True,
        secure=config.COOKIE_SECURE,
        samesite=config.COOKIE_SAMESITE,
        max_age=config.COOKIE_MAX_AGE,
        path="/",
    )


@app.get("/api/health")
def health_check():
    from paths import resolve_matriz_cursos_path
    return {
        "status": "ok",
        "env": config.APP_ENV,
        "database": "configured" if config.SQLALCHEMY_DATABASE_URL else "missing",
        "matriz_cursos": bool(resolve_matriz_cursos_path()),
    }


@app.post("/api/auth/login")
def login(
    request: Request,
    response: Response,
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db),
):
    check_login_rate_limit(request)
    username = (form_data.username or "").strip()
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user:
        user = (
            db.query(models.User)
            .filter(models.User.username.ilike(username))
            .first()
        )
    try:
        password_ok = user and auth.verify_password(form_data.password, user.hashed_password)
    except Exception:
        logger.exception("Error verificando contraseña para usuario %s", username)
        password_ok = False
    if not password_ok:
        raise HTTPException(status_code=400, detail="Usuario o contraseña incorrectos")

    access_token = auth.create_access_token(data={"sub": user.username})
    set_auth_cookie(response, access_token)
    return {"token_type": "bearer", "message": "ok"}


@app.post("/api/auth/logout")
def logout(response: Response):
    response.delete_cookie(
        key=config.COOKIE_NAME,
        path="/",
        secure=config.COOKIE_SECURE,
        samesite=config.COOKIE_SAMESITE,
    )
    return {"message": "ok"}

@app.get("/api/auth/me")
def read_users_me(current_user: models.User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "username": current_user.username,
        "role": current_user.role,
        "permissions": json.loads(current_user.permissions_json)
    }

class UserCreate(BaseModel):
    username: str
    password: str
    role: str
    permissions: list[str]

@app.get("/api/users")
def get_users(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Not superadmin")
    users = db.query(models.User).all()
    return [{"id": u.id, "username": u.username, "role": u.role, "permissions": json.loads(u.permissions_json)} for u in users]

@app.post("/api/users")
def create_user(user: UserCreate, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Not superadmin")
    if db.query(models.User).filter(models.User.username == user.username).first():
        raise HTTPException(status_code=400, detail="Username already registered")
    if user.role == "superadmin" and current_user.username != config.ADMIN_MASTER_USER:
        raise HTTPException(status_code=403, detail="Solo el administrador principal puede crear superadmins")

    validate_role(user.role)
    perms = validate_permissions(user.permissions)
    if len(user.password) < 8:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")

    hashed_password = auth.get_password_hash(user.password)
    db_user = models.User(
        username=user.username,
        hashed_password=hashed_password,
        role=user.role,
        permissions_json=json.dumps(perms)
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return {"id": db_user.id, "username": db_user.username}

@app.put("/api/users/{user_id}")
def update_user(user_id: int, user: UserCreate, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Not superadmin")
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.role == "superadmin" and current_user.username != config.ADMIN_MASTER_USER and db_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Solo el administrador principal puede asignar rol superadmin")

    validate_role(user.role)
    perms = validate_permissions(user.permissions)
    db_user.username = user.username
    if user.password:
        if len(user.password) < 8:
            raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")
        db_user.hashed_password = auth.get_password_hash(user.password)
    db_user.role = user.role
    db_user.permissions_json = json.dumps(perms)
    db.commit()
    return {"status": "updated"}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Not superadmin")
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    if db_user.username == "admin":
        raise HTTPException(status_code=400, detail="Cannot delete default admin")
    db.delete(db_user)
    db.commit()
    return {"status": "deleted"}

# === RUTAS PARA NORMALIZADOR DE RUT ===
class RutInput(BaseModel):
    ruts: str
    formato: str
    k_minuscula: bool

@app.post("/api/rut/normalizar")
def normalizar_rut(data: RutInput, current_user: models.User = Depends(require_permission("rut"))):
    ruts_limpios = data.ruts.replace("'", "").replace('"', "").replace('\r', '')
    lista_ruts = [r.strip() for r in ruts_limpios.split('\n') if r.strip()]
    
    resultados = []
    for rut in lista_ruts:
        rut_base = rut.upper().replace(".", "")
        if "-" not in rut_base and len(rut_base) > 1:
            rut_base = rut_base[:-1] + "-" + rut_base[-1]
            
        partes = rut_base.split("-")
        if len(partes) != 2:
            resultados.append(rut)
            continue
            
        cuerpo, dv = partes[0], partes[1]
        if data.k_minuscula and dv == 'K': dv = 'k'
            
        if data.formato == "Sin Puntos y Con Guión":
            resultados.append(f"{cuerpo}-{dv}")
        elif data.formato == "Sin Puntos y Sin Guión":
            resultados.append(f"{cuerpo}{dv}")
        else:
            cuerpo_puntos = f"{int(cuerpo):,}".replace(",", ".") if cuerpo.isdigit() else cuerpo
            resultados.append(f"{cuerpo_puntos}-{dv}")
            
    return {"ruts": "\n".join(resultados), "total": len(resultados)}


# === RUTAS PARA NORMALIZADOR DE NOMBRES ===
class NombresInput(BaseModel):
    nombres: str
    formato: str

def _normalizar_lista_textos(nombres: str, formato: str) -> dict:
    nombres_limpios = nombres.replace("'", "").replace('"', "").replace("\r", "")
    lista_nombres = [n.strip() for n in nombres_limpios.split("\n") if n.strip()]

    resultados = []
    for nom in lista_nombres:
        if formato == "Mayúsculas":
            resultados.append(nom.upper())
        elif formato == "Minúsculas":
            resultados.append(nom.lower())
        elif formato in ("Primera Letra Mayúscula", "Título", "Title"):
            resultados.append(nom.title())
        else:
            resultados.append(nom.strip())

    return {"nombres": "\n".join(resultados), "total": len(resultados)}


@app.post("/api/textos/normalizar")
def normalizar_textos(data: NombresInput, current_user: models.User = Depends(require_permission("textos"))):
    return _normalizar_lista_textos(data.nombres, data.formato)


@app.post("/api/nombres/normalizar")
def normalizar_nombres(data: NombresInput, current_user: models.User = Depends(require_permission("textos"))):
    return _normalizar_lista_textos(data.nombres, data.formato)


# === REPORTE CONSULTA CURSOS (MOODLE) ===
from consulta_cursos_parser import build_excel_bytes, parse_consulta_text, summarize_rows


class ConsultaCursosInput(BaseModel):
    texto: str


@app.post("/api/consulta-cursos/preview")
def preview_consulta_cursos(
    data: ConsultaCursosInput,
    current_user: models.User = Depends(require_permission("consulta_cursos")),
):
    texto = (data.texto or "").strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Pegue el resultado de la consulta Moodle.")
    rows = parse_consulta_text(texto)
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron registros. Verifique que el texto incluya líneas «Curso: … (ID …)» y filas con RUT.",
        )
    summary = summarize_rows(rows)
    return {
        **summary,
        "muestra": [r.as_data_list() for r in rows[:5]],
    }


@app.post("/api/consulta-cursos/excel")
def export_consulta_cursos_excel(
    data: ConsultaCursosInput,
    current_user: models.User = Depends(require_permission("consulta_cursos")),
):
    texto = (data.texto or "").strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Pegue el resultado de la consulta Moodle.")
    rows = parse_consulta_text(texto)
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron registros para exportar.",
        )
    try:
        content, filename = build_excel_bytes(rows)
        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=generic_error_detail(e, "reporte consulta cursos"))


# === USUARIOS DUPLICADOS (mdl_user) ===
from mdl_user_duplicados import (
    analyze_duplicates,
    build_excel_bytes as build_duplicados_excel,
    build_result_payload,
    read_mdl_user_file,
)

MDL_USER_HISTORY_KEY = "usuarios_duplicados"
MDL_USER_MAX_SCANS = 20


def _load_duplicados_history(username: str, db: Session) -> dict:
    data = get_json_db(MDL_USER_HISTORY_KEY, username, db)
    if isinstance(data, dict) and isinstance(data.get("scans"), list):
        return data
    return {"scans": []}


def _save_duplicados_scan(username: str, scan: dict, db: Session) -> dict:
    history = _load_duplicados_history(username, db)
    scans = history.get("scans", [])
    scans.insert(0, scan)
    history["scans"] = scans[:MDL_USER_MAX_SCANS]
    save_db(MDL_USER_HISTORY_KEY, username, history, db)
    return scan


@app.post("/api/usuarios-duplicados/analizar")
async def analizar_usuarios_duplicados(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_permission("usuarios_duplicados")),
    db: Session = Depends(get_db),
):
    path = None
    try:
        safe_name = safe_planilla_filename(file.filename)
        content = await read_upload_limited(file)
        path = os.path.join(tempfile.mkdtemp(), safe_name)
        with open(path, "wb") as f:
            f.write(content)

        _, rows = read_mdl_user_file(path)
        analysis = analyze_duplicates(rows)
        scan = build_result_payload(file.filename or safe_name, analysis)
        _save_duplicados_scan(current_user.username, scan, db)
        return scan
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=generic_error_detail(exc, "análisis de duplicados"))
    finally:
        if path and os.path.exists(path):
            shutil.rmtree(os.path.dirname(path), ignore_errors=True)


@app.get("/api/usuarios-duplicados/historial")
def historial_usuarios_duplicados(
    current_user: models.User = Depends(require_permission("usuarios_duplicados")),
    db: Session = Depends(get_db),
):
    history = _load_duplicados_history(current_user.username, db)
    summaries = []
    for scan in history.get("scans", []):
        summaries.append(
            {
                "id": scan.get("id"),
                "filename": scan.get("filename"),
                "analyzed_at": scan.get("analyzed_at"),
                "total_rows": scan.get("total_rows", 0),
                "duplicate_groups": scan.get("duplicate_groups", 0),
                "duplicate_rows": scan.get("duplicate_rows", 0),
                "rows_without_key": scan.get("rows_without_key", 0),
                "by_criterion": scan.get("by_criterion", {}),
            }
        )
    return {"scans": summaries}


@app.get("/api/usuarios-duplicados/historial/{scan_id}")
def obtener_scan_usuarios_duplicados(
    scan_id: str,
    current_user: models.User = Depends(require_permission("usuarios_duplicados")),
    db: Session = Depends(get_db),
):
    history = _load_duplicados_history(current_user.username, db)
    for scan in history.get("scans", []):
        if scan.get("id") == scan_id:
            return scan
    raise HTTPException(status_code=404, detail="Análisis no encontrado.")


@app.delete("/api/usuarios-duplicados/historial")
def limpiar_historial_usuarios_duplicados(
    current_user: models.User = Depends(require_permission("usuarios_duplicados")),
    db: Session = Depends(get_db),
):
    save_db(MDL_USER_HISTORY_KEY, current_user.username, {"scans": []}, db)
    return {"status": "success"}


class DuplicadosExcelInput(BaseModel):
    scan_id: str | None = None


@app.post("/api/usuarios-duplicados/excel")
def exportar_usuarios_duplicados_excel(
    data: DuplicadosExcelInput,
    current_user: models.User = Depends(require_permission("usuarios_duplicados")),
    db: Session = Depends(get_db),
):
    if not data.scan_id:
        raise HTTPException(status_code=400, detail="Indique el análisis a exportar.")
    history = _load_duplicados_history(current_user.username, db)
    scan = next((s for s in history.get("scans", []) if s.get("id") == data.scan_id), None)
    if not scan:
        raise HTTPException(status_code=404, detail="Análisis no encontrado.")
    try:
        content, filename = build_duplicados_excel(scan)
        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=generic_error_detail(exc, "exportación duplicados"))


# === CRUD JSON ===
def get_json_db(db_name: str, username: str, db_session: Session):
    record = db_session.query(models.AppData).filter(
        models.AppData.username == username,
        models.AppData.module_name == db_name
    ).first()
    
    if record and record.payload_json:
        try:
            data = json.loads(record.payload_json)
            if db_name == "smtp_config":
                return expose_smtp_config(data, include_secret=True)
            return data
        except:
            pass
    return {} if db_name == "recordatorios" else []

def save_db(db_name: str, username: str, data, db_session: Session):
    to_store = data
    if db_name == "smtp_config" and isinstance(data, dict):
        record = db_session.query(models.AppData).filter(
            models.AppData.username == username,
            models.AppData.module_name == db_name,
        ).first()
        existing_raw = {}
        if record and record.payload_json:
            try:
                existing_raw = json.loads(record.payload_json)
            except json.JSONDecodeError:
                existing_raw = {}
        merged = dict(data)
        if not merged.get("password") and existing_raw.get("password"):
            merged["password"] = existing_raw["password"]
        to_store = protect_smtp_config(merged)
    payload = json.dumps(to_store)
    record = db_session.query(models.AppData).filter(
        models.AppData.username == username,
        models.AppData.module_name == db_name
    ).first()
    
    if record:
        record.payload_json = payload
    else:
        new_record = models.AppData(username=username, module_name=db_name, payload_json=payload)
        db_session.add(new_record)
    db_session.commit()

@app.get("/api/db/{db_name}")
def read_db(db_name: str, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "superadmin":
        user_permissions = json.loads(current_user.permissions_json)
        if db_name not in user_permissions and not (db_name == "smtp_config" and "recordatorios" in user_permissions):
            raise HTTPException(status_code=403, detail="Not enough permissions")
    if db_name not in config.VALID_DB_MODULES:
        raise HTTPException(status_code=404, detail="DB not found")
    return get_json_db(db_name, current_user.username, db)

@app.post("/api/db/{db_name}")
def write_db(db_name: str, data: dict | list = Body(...), current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "superadmin":
        user_permissions = json.loads(current_user.permissions_json)
        if db_name not in user_permissions and not (db_name == "smtp_config" and "recordatorios" in user_permissions):
            raise HTTPException(status_code=403, detail="Not enough permissions")
    if db_name not in config.VALID_DB_MODULES:
        raise HTTPException(status_code=404, detail="DB not found")
    save_db(db_name, current_user.username, data, db)
    return {"status": "success"}

# === COMPARADOR DE DATOS ===
def column_index_from_string(col):
    col = col.upper()
    num = 0
    for c in col:
        num = num * 26 + (ord(c) - ord('A')) + 1
    return num

def extraer_excel(path, col_ini_letra, col_fin_letra, fila_ini, fila_fin, hoja_name=None):
    usuarios = set()
    row_map = {}
    col_num_1 = column_index_from_string(col_ini_letra)
    col_num_2 = column_index_from_string(col_fin_letra) if col_fin_letra else col_num_1
    c_start, c_end = min(col_num_1, col_num_2), max(col_num_1, col_num_2)
    
    wb = openpyxl.load_workbook(path, data_only=True)
    if hoja_name and hoja_name != "Activa (Por defecto)" and hoja_name in wb.sheetnames:
        sheet = wb[hoja_name]
    else:
        sheet = wb.active
    max_row = fila_fin if fila_fin else sheet.max_row
    
    for row in sheet.iter_rows(min_row=fila_ini, max_row=max_row, min_col=1, max_col=sheet.max_column, values_only=True):
        for c_idx in range(c_start - 1, c_end):
            if c_idx < len(row):
                val = row[c_idx]
                if val is not None and str(val).strip() != "":
                    clean_val = str(val).strip().lower()
                    usuarios.add(clean_val)
                    if clean_val not in row_map:
                        row_map[clean_val] = list(row)
    return usuarios, row_map

@app.post("/api/excel/hojas")
async def get_excel_hojas(file: UploadFile = File(...), current_user: models.User = Depends(require_permission("comparador"))):
    path = None
    try:
        safe_name = safe_upload_filename(file.filename)
        content = await read_upload_limited(file)
        path = os.path.join(tempfile.mkdtemp(), safe_name)
        with open(path, "wb") as f:
            f.write(content)
        wb = openpyxl.load_workbook(path, read_only=True)
        hojas = wb.sheetnames
        wb.close()
        return {"hojas": hojas}
    except HTTPException:
        raise
    except Exception:
        return {"hojas": []}
    finally:
        if path and os.path.exists(path):
            shutil.rmtree(os.path.dirname(path), ignore_errors=True)



@app.post("/api/comparador")
async def api_comparar(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    tipo_reporte: str = Form(...),
    c_ini1: str = Form("A"),
    c_fin1: str = Form(""),
    f_ini1: int = Form(2),
    f_fin1: str = Form(""),
    hoja1: str = Form("Activa (Por defecto)"),
    c_ini2: str = Form("A"),
    c_fin2: str = Form(""),
    f_ini2: int = Form(2),
    f_fin2: str = Form(""),
    hoja2: str = Form("Activa (Por defecto)"),
    current_user: models.User = Depends(require_permission("comparador"))
):
    path1 = path2 = None
    temp_cmp = None
    try:
        temp_cmp = tempfile.mkdtemp()
        safe1 = safe_upload_filename(file1.filename)
        safe2 = safe_upload_filename(file2.filename)
        path1 = os.path.join(temp_cmp, safe1)
        path2 = os.path.join(temp_cmp, safe2)

        with open(path1, "wb") as f:
            f.write(await read_upload_limited(file1))
        with open(path2, "wb") as f:
            f.write(await read_upload_limited(file2))
            
        f_fin1_int = int(f_fin1) if f_fin1.isdigit() else None
        f_fin2_int = int(f_fin2) if f_fin2.isdigit() else None
        
        datos1, rows1 = extraer_excel(path1, c_ini1, c_fin1 if c_fin1 else None, f_ini1, f_fin1_int, hoja1)
        datos2, rows2 = extraer_excel(path2, c_ini2, c_fin2 if c_fin2 else None, f_ini2, f_fin2_int, hoja2)
        
        coincidencias = datos1.intersection(datos2)
        solo_1 = datos1 - datos2
        solo_2 = datos2 - datos1
        
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        hay_diferencias = len(solo_1) + len(solo_2) > 0
        hay_coincidencias = len(coincidencias) > 0

        if tipo_reporte in ["diferencias", "ambos"] and hay_diferencias:
            ws_diff = wb_out.create_sheet("Diferencias")
            max_len = 0
            for d in solo_1: max_len = max(max_len, len(rows1.get(d, [])))
            for d in solo_2: max_len = max(max_len, len(rows2.get(d, [])))
            header_row = [f"Columna {i+1}" for i in range(max_len)] + ["Estado (Diferencia)"]
            ws_diff.append(header_row)
            
            for dato in solo_1:
                row_data = list(rows1.get(dato, [dato]))
                while len(row_data) < max_len: row_data.append("")
                row_data.append(f"Solo en {file1.filename}")
                ws_diff.append(row_data)
            for dato in solo_2:
                row_data = list(rows2.get(dato, [dato]))
                while len(row_data) < max_len: row_data.append("")
                row_data.append(f"Solo en {file2.filename}")
                ws_diff.append(row_data)

        if tipo_reporte in ["coincidencias", "ambos"] and hay_coincidencias:
            ws_coinc = wb_out.create_sheet("Coincidencias")
            max_len1 = max([len(rows1.get(d, [])) for d in coincidencias] + [0])
            max_len2 = max([len(rows2.get(d, [])) for d in coincidencias] + [0])
            header_row = [f"Base Col {i+1}" for i in range(max_len1)] + ["|| MATCH ||"] + [f"Contraste Col {i+1}" for i in range(max_len2)]
            ws_coinc.append(header_row)
            
            for dato in coincidencias:
                r1 = list(rows1.get(dato, [dato]))
                while len(r1) < max_len1: r1.append("")
                r2 = list(rows2.get(dato, [dato]))
                while len(r2) < max_len2: r2.append("")
                ws_coinc.append(r1 + ["<--- SI --->"] + r2)

        if not wb_out.sheetnames:
            ws_empty = wb_out.create_sheet("Sin Datos")
            ws_empty.append(["No hubo datos para el reporte seleccionado."])

        out_name = f"Reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb_out.save(out_name)
        
        if temp_cmp:
            shutil.rmtree(temp_cmp, ignore_errors=True)

        return FileResponse(path=out_name, filename=out_name, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=generic_error_detail(e, "comparación"))
    finally:
        if temp_cmp and os.path.isdir(temp_cmp):
            shutil.rmtree(temp_cmp, ignore_errors=True)


# === GENERADOR DE CARGAS / SCRIPTS DE INDUCCIÓN ===
from matriz_cursos import extract_perfil_from_row
from matriz_db import (
    courses_for_perfil as db_courses_for_perfil,
    create_profile,
    delete_profile,
    ensure_matriz_seeded,
    get_matriz_info_db,
    list_courses,
    list_profiles,
    seed_profiles_from_excel,
    sync_catalog_from_excel,
    update_profile,
)


class ProfileCreateBody(BaseModel):
    name: str
    course_moodle_ids: list[int] = []


class ProfileUpdateBody(BaseModel):
    name: str | None = None
    course_moodle_ids: list[int] | None = None


def normalize_text(val):
    if val is None:
        return ""
    s = str(val).strip()
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.upper()

def clean_username_rut(rut_val):
    if not rut_val:
        return ""
    r = str(rut_val).strip().replace(".", "").replace("-", "").replace(" ", "")
    if not r:
        return ""
    if r[-1].upper() == 'K':
        r = r[:-1] + 'k'
    return r

def process_row_mapping(row, col_name_to_idx, mapping, grupo, db: Session):
    processed_row = {}
    
    for out_col, cfg in mapping.items():
        val_type = cfg.get("type")
        val_setting = cfg.get("value")
        
        raw_val = ""
        if val_type == "fixed":
            raw_val = val_setting
        elif val_type == "column":
            idx = col_name_to_idx.get(val_setting)
            if idx is not None and idx < len(row):
                raw_val = row[idx]
        
        # Apply normalizations
        if out_col in ["username", "password"]:
            processed_row[out_col] = clean_username_rut(raw_val)
        elif out_col in ["firstname", "lastname"]:
            processed_row[out_col] = str(raw_val).upper().strip() if raw_val is not None else ""
        elif out_col == "email":
            processed_row[out_col] = str(raw_val).upper().strip() if raw_val is not None else ""
        elif out_col == "department":
            processed_row[out_col] = str(raw_val).upper().strip() if raw_val is not None else ""
        else:
            processed_row[out_col] = str(raw_val).strip() if raw_val is not None else ""
            
    department_val = processed_row.get("department", "")
    perfil_original, perfil_norm = extract_perfil_from_row(
        row,
        col_name_to_idx,
        processed_department=department_val,
    )
    if not perfil_original:
        perfil_original = "SIN PERFIL"

    courses = db_courses_for_perfil(db, perfil_norm, perfil_original)
    
    return {
        "processed_row": processed_row,
        "perfil_norm": perfil_norm,
        "perfil_original": perfil_original,
        "courses": courses
    }

@app.get("/api/excel/matriz-info")
def api_matriz_info(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    return get_matriz_info_db(db)


@app.get("/api/generador/cursos")
def api_list_cursos(
    search: str = "",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    return {"cursos": list_courses(db, search=search)}


@app.get("/api/generador/perfiles")
def api_list_perfiles(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    return {"perfiles": list_profiles(db)}


@app.post("/api/generador/perfiles")
def api_create_perfil(
    body: ProfileCreateBody,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    try:
        profile = create_profile(db, body.name, body.course_moodle_ids)
        return profile
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/generador/perfiles/{profile_id}")
def api_update_perfil(
    profile_id: int,
    body: ProfileUpdateBody,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    try:
        return update_profile(db, profile_id, body.name, body.course_moodle_ids)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/generador/perfiles/{profile_id}")
def api_delete_perfil(
    profile_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    try:
        delete_profile(db, profile_id)
        return {"ok": True}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/api/generador/sync-catalogo")
def api_sync_catalogo(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    n = sync_catalog_from_excel(db)
    if n == 0:
        raise HTTPException(
            status_code=404,
            detail="No se encontró «cursos bex Moodle» para importar el catálogo.",
        )
    return {"importados": n}


@app.post("/api/generador/reimportar-perfiles-excel")
def api_reimportar_perfiles(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Solo superadmin")
    n = seed_profiles_from_excel(db, replace_existing=True)
    return {"perfiles": n}


@app.post("/api/excel/inspect")
async def api_inspect_excel(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_permission("generador")),
):
    try:
        temp_dir = tempfile.mkdtemp()
        safe_name = safe_upload_filename(file.filename)
        path = os.path.join(temp_dir, safe_name)
        with open(path, "wb") as f:
            f.write(await read_upload_limited(file))
            
        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = {}
        for sname in wb.sheetnames:
            sheet = wb[sname]
            headers = []
            for row in sheet.iter_rows(max_row=1, values_only=True):
                headers = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                break
            sheets[sname] = headers
            
        wb.close()
        shutil.rmtree(temp_dir)
        return {"sheets": sheets}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=generic_error_detail(e, "inspección del archivo"))

@app.post("/api/excel/preview")
async def api_preview_carga(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    grupo: str = Form(...),
    mapping: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    try:
        mapping_dict = json.loads(mapping)
        ensure_matriz_seeded(db)

        temp_dir = tempfile.mkdtemp()
        safe_name = safe_upload_filename(file.filename)
        path = os.path.join(temp_dir, safe_name)
        with open(path, "wb") as f:
            f.write(await read_upload_limited(file))

        wb_dot = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb_dot.sheetnames:
            raise HTTPException(status_code=400, detail=f"No se encontró la hoja '{sheet_name}' en el Excel.")
            
        sheet_dot = wb_dot[sheet_name]
        rows_dot = list(sheet_dot.iter_rows(values_only=True))
        wb_dot.close()
        shutil.rmtree(temp_dir)
        
        if not rows_dot:
            raise HTTPException(status_code=400, detail="La hoja seleccionada está vacía.")
            
        header = [str(c).strip() if c else "" for c in rows_dot[0]]
        col_name_to_idx = {name: idx for idx, name in enumerate(header) if name}
        
        previews = {}
        data_rows = rows_dot[1:]
        perfiles_sin_cursos: set[str] = set()
        
        for r in data_rows:
            if not any(r):
                continue
                
            res = process_row_mapping(r, col_name_to_idx, mapping_dict, grupo, db)
            if not res["courses"] and res["perfil_norm"] != "SIN_PERFIL":
                perfiles_sin_cursos.add(res["perfil_original"])
            p_norm = res["perfil_norm"]
            p_original = res["perfil_original"]
            processed_row = res["processed_row"]
            courses = res["courses"]
            
            if p_norm not in previews:
                out_headers = list(mapping_dict.keys())
                for i in range(1, len(courses) + 1):
                    out_headers.append(f"group{i}")
                    out_headers.append(f"course{i}")
                
                previews[p_norm] = {
                    "profile_name": p_original,
                    "headers": out_headers,
                    "rows": []
                }
                
            if len(previews[p_norm]["rows"]) < 10:
                row_vals = []
                for out_col in mapping_dict.keys():
                    row_vals.append(processed_row.get(out_col, ""))
                for course in courses:
                    row_vals.append(grupo)
                    row_vals.append(course)
                previews[p_norm]["rows"].append(row_vals)

        warnings = []
        if perfiles_sin_cursos:
            info = get_matriz_info_db(db)
            nombres = ", ".join(p["hoja"] for p in info.get("perfiles", []))
            warnings.append(
                "Perfiles sin cursos asignados: "
                + ", ".join(sorted(perfiles_sin_cursos))
                + f". Cree o edite el perfil en «Perfiles de inducción» (disponibles: {nombres})."
                + " El valor de PERFIL DE INDUCCIÓN debe coincidir con el nombre del perfil."
            )
        if not previews:
            warnings.append(
                "No se generaron filas. Verifique usuario/RUT y que exista columna PERFIL DE INDUCCIÓN."
            )

        return {"previews": previews, "warnings": warnings}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=generic_error_detail(e, "previsualización"))

@app.post("/api/excel/generar-carga")
async def api_generar_carga(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    grupo: str = Form(...),
    mapping: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("generador")),
):
    try:
        mapping_dict = json.loads(mapping)
        ensure_matriz_seeded(db)

        temp_dir = tempfile.mkdtemp()
        safe_name = safe_upload_filename(file.filename)
        path = os.path.join(temp_dir, safe_name)
        with open(path, "wb") as f:
            f.write(await read_upload_limited(file))

        wb_dot = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb_dot.sheetnames:
            raise HTTPException(status_code=400, detail=f"No se encontró la hoja '{sheet_name}' en el Excel.")
            
        sheet_dot = wb_dot[sheet_name]
        rows_dot = list(sheet_dot.iter_rows(values_only=True))
        wb_dot.close()
        
        if not rows_dot:
            raise HTTPException(status_code=400, detail="El archivo de dotación está vacío.")
            
        header = [str(c).strip() if c else "" for c in rows_dot[0]]
        col_name_to_idx = {name: idx for idx, name in enumerate(header) if name}
        
        colaboradores_por_perfil = {}
        for r_idx, r in enumerate(rows_dot[1:]):
            if not any(r):
                continue
                
            res = process_row_mapping(r, col_name_to_idx, mapping_dict, grupo, db)
            p_norm = res["perfil_norm"]
            p_original = res["perfil_original"]
            processed_row = res["processed_row"]
            courses = res["courses"]
            
            username = processed_row.get("username")
            if not username:
                continue
                
            if p_norm not in colaboradores_por_perfil:
                colaboradores_por_perfil[p_norm] = {
                    "original_name": p_original,
                    "courses": courses,
                    "items": []
                }
            colaboradores_por_perfil[p_norm]["items"].append(processed_row)
            
        generated_files = []
        for p_norm, p_data in colaboradores_por_perfil.items():
            original_name = p_data["original_name"]
            items = p_data["items"]
            cursos = p_data["courses"]
            
            out_headers = list(mapping_dict.keys())
            for i in range(1, len(cursos) + 1):
                out_headers.append(f"group{i}")
                out_headers.append(f"course{i}")
                
            safe_p_name = "".join(c for c in p_norm if c.isalnum() or c in (" ", "_", "-")).replace(" ", "_")
            csv_filename = f"script_{safe_p_name}.csv"
            csv_filepath = os.path.join(temp_dir, csv_filename)
            
            with open(csv_filepath, mode="w", newline="", encoding="utf-8-sig") as csv_file:
                writer = csv.writer(csv_file, delimiter=";")
                writer.writerow(out_headers)
                
                for item in items:
                    row = []
                    for out_col in mapping_dict.keys():
                        row.append(item.get(out_col, ""))
                    for course in cursos:
                        row.append(grupo)
                        row.append(course)
                    writer.writerow(row)
                    
            generated_files.append(csv_filepath)
            
        if not generated_files:
            raise HTTPException(status_code=400, detail="No se generaron registros válidos del archivo de entrada.")
            
        zip_name = f"Cargas_Induccion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = os.path.join(temp_dir, zip_name)
        
        with zipfile.ZipFile(zip_path, "w") as zip_f:
            for fpath in generated_files:
                zip_f.write(fpath, os.path.basename(fpath))
                
        dest_zip_path = os.path.abspath(zip_name)
        shutil.copy(zip_path, dest_zip_path)
        
        shutil.rmtree(temp_dir)
        
        background_tasks.add_task(os.remove, dest_zip_path)
        return FileResponse(path=dest_zip_path, filename=zip_name, media_type="application/zip")
        
    except HTTPException as he:
        raise he
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=generic_error_detail(e, "generación de cargas"))


from transelec_routes import router as transelec_router
from aza_routes import router as aza_router
from resiter_routes import router as resiter_router
from carozzi_routes import router as carozzi_router

app.include_router(transelec_router)
app.include_router(aza_router)
app.include_router(resiter_router)
app.include_router(carozzi_router)

from compresor_routes import router as compresor_router

app.include_router(compresor_router)

# === GENERADOR DE PAQUETES DE VIDEO ===
from pathlib import Path as FsPath
from video_packages import (
    crear_zip_lote,
    generar_paquetes_video,
    validar_nombre_carpeta,
    validar_url_curso,
)


@app.post("/api/generador/videos/generar")
async def api_generar_paquetes_video(
    background_tasks: BackgroundTasks,
    nombre_lote: str = Form(...),
    course_url: str = Form(...),
    videos: list[UploadFile] = File(...),
    current_user: models.User = Depends(require_permission("generador")),
):
    error = validar_nombre_carpeta(nombre_lote)
    if error:
        raise HTTPException(status_code=400, detail=error)

    error = validar_url_curso(course_url)
    if error:
        raise HTTPException(status_code=400, detail=error)

    if not videos:
        raise HTTPException(status_code=400, detail="Agrega al menos un video.")

    temp_dir = tempfile.mkdtemp()
    saved_paths: list[FsPath] = []
    total_bytes = 0

    try:
        for upload in videos:
            safe_name = safe_video_filename(upload.filename)
            dest = FsPath(temp_dir) / f"{len(saved_paths)}_{safe_name}"
            file_size = await stream_video_upload_to_file(upload, dest)
            total_bytes += file_size
            if total_bytes > config.MAX_VIDEO_BATCH_BYTES:
                raise HTTPException(
                    status_code=413,
                    detail=(
                        f"El total de videos supera el límite de "
                        f"{config.MAX_VIDEO_BATCH_BYTES // (1024 * 1024)} MB."
                    ),
                )
            saved_paths.append(dest)

        lote_dir = generar_paquetes_video(
            FsPath(temp_dir),
            nombre_lote.strip(),
            course_url.strip(),
            saved_paths,
        )

        zip_name = f"{nombre_lote.strip()}.zip"
        zip_path = FsPath(temp_dir) / zip_name
        crear_zip_lote(lote_dir, zip_path)

        dest_zip = os.path.abspath(zip_name)
        shutil.copy(zip_path, dest_zip)
        shutil.rmtree(temp_dir, ignore_errors=True)

        background_tasks.add_task(os.remove, dest_zip)
        return FileResponse(dest_zip, filename=zip_name, media_type="application/zip")
    except HTTPException:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise HTTPException(
            status_code=500,
            detail=generic_error_detail(e, "generación de paquetes de video"),
        )

