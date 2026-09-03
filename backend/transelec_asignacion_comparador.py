"""Comparador Asignación RH vs Nómina LMS (Transelec).

Verifica que cada persona del Excel de Asignación exista en el CSV de Nómina
y que email / RUT / nombre correspondan a la misma persona.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COLS_FALTANTES = [
    "Periodo",
    "RUT",
    "Digito",
    "Nombre",
    "EMAIL",
    "Cargo",
    "Unidad",
    "Gcia/SubGcia",
    "VP",
    "Ubicacion",
    "Nombre Jefe",
    "TIPO CONTRATO",
    "CeCo",
]

COLS_INCONSISTENCIAS = [
    "Motivo",
    "Criterio match",
    "Asignacion EMAIL",
    "Asignacion RUT",
    "Asignacion Nombre",
    "Nomina username",
    "Nomina email",
    "Nomina RUT",
    "Nomina Nombre",
    "Nomina suspended",
    "Detalle",
]


def normalizar_email(valor: Any) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    return str(valor).strip().lower()


def normalizar_rut(valor: Any) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    s = str(valor).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace("\xa0", "").replace(" ", "")
    return re.sub(r"[^0-9Kk]", "", s).upper()


def normalizar_nombre(valor: Any) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    s = str(valor).strip().casefold()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _tokens_nombre(valor: str) -> set[str]:
    return {t for t in normalizar_nombre(valor).split() if len(t) > 1}


def nombres_compatibles(nombre_asig: str, nombre_nomina: str) -> bool:
    """True si hay solapamiento razonable de tokens (evita falsos positivos por orden)."""
    a = _tokens_nombre(nombre_asig)
    b = _tokens_nombre(nombre_nomina)
    if not a or not b:
        return True  # sin dato suficiente para afirmar conflicto
    comunes = a & b
    if len(comunes) >= 2:
        return True
    if len(comunes) == 1 and (len(a) == 1 or len(b) == 1):
        return True
    # Si hay al menos un apellido/nombre largo en común
    if any(len(t) >= 5 for t in comunes):
        return True
    return False


def ruts_coinciden(rut_asig: str, rut_asig_dv: str, rut_nomina: str) -> bool | None:
    """
    True = mismo RUT, False = distinto, None = no se puede comparar (falta dato).
    Acepta RUT con o sin dígito verificador en nómina.
    """
    if not rut_nomina or rut_nomina == "-":
        return None
    candidatos = {r for r in (rut_asig, rut_asig_dv) if r}
    if not candidatos:
        return None
    if rut_nomina in candidatos:
        return True
    # Comparar sin DV: si nómina trae cuerpo y asignación tiene cuerpo
    if rut_asig and (
        rut_nomina == rut_asig
        or rut_nomina.startswith(rut_asig)
        or rut_asig.startswith(rut_nomina)
    ):
        # Evitar matches demasiado cortos
        if len(rut_asig) >= 6 and len(rut_nomina) >= 6:
            shorter, longer = sorted([rut_asig, rut_nomina], key=len)
            if longer.startswith(shorter) and len(longer) - len(shorter) <= 1:
                return True
    return False


def _detectar_encoding(ruta: Path | str) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(ruta, encoding=encoding) as f:
                f.readline()
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detectar_separador(texto: str) -> str:
    primera = texto.splitlines()[0] if texto else ""
    try:
        dialecto = csv.Sniffer().sniff(texto[:8192], delimiters=";,")
        if dialecto.delimiter in (";", ","):
            return dialecto.delimiter
    except csv.Error:
        pass
    return ";" if primera.count(";") >= primera.count(",") else ","


def _leer_csv(ruta: Path | str) -> pd.DataFrame:
    encoding = _detectar_encoding(ruta)
    with open(ruta, encoding=encoding, errors="replace") as f:
        texto = f.read()
    sep = _detectar_separador(texto)
    kwargs = dict(sep=sep, dtype=str, engine="python", quotechar='"', skipinitialspace=True)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=pd.errors.ParserWarning)
        try:
            return pd.read_csv(io.StringIO(texto), on_bad_lines="warn", **kwargs)
        except TypeError:
            return pd.read_csv(
                io.StringIO(texto),
                error_bad_lines=False,
                warn_bad_lines=False,
                **kwargs,
            )


def _buscar_columna(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    mapa = {str(c).strip().casefold(): c for c in df.columns}
    for cand in candidatos:
        key = cand.strip().casefold()
        if key in mapa:
            return mapa[key]
    for col in df.columns:
        limpio = re.sub(r"[^a-z0-9]", "", str(col).casefold())
        for cand in candidatos:
            cand_limpo = re.sub(r"[^a-z0-9]", "", cand.casefold())
            if limpio == cand_limpo:
                return col
            if cand_limpo in ("email", "rut", "nombre", "username", "firstname", "lastname", "address", "idnumber"):
                if limpio == cand_limpo or limpio.endswith(cand_limpo):
                    return col
    return None


def _buscar_columna_digito(df: pd.DataFrame) -> str | None:
    for col in df.columns:
        limpio = re.sub(r"[^a-z0-9]", "", str(col).casefold())
        if limpio in ("digito", "dv", "digitoverificador") or "gito" in limpio:
            return col
    return None


def _rut_desde_fila(row: dict[str, str], col_addr: str | None, col_id: str | None) -> str:
    for col in (col_addr, col_id):
        if not col:
            continue
        rut = normalizar_rut(row.get(col))
        if rut and rut != "-":
            return rut
    return ""


def leer_nomina(ruta: Path | str) -> dict[str, Any]:
    """Índices de búsqueda con registros completos de la nómina LMS."""
    path = Path(ruta)
    if path.suffix.lower() != ".csv":
        raise ValueError("El archivo de Nómina debe ser CSV (.csv).")

    df = _leer_csv(path)
    col_email = _buscar_columna(df, ["email", "Email", "correo"])
    col_user = _buscar_columna(df, ["username", "Username", "user"])
    col_addr = _buscar_columna(df, ["address", "Address", "direccion"])
    col_id = _buscar_columna(df, ["idnumber", "id_number", "RUT", "rut"])
    col_fn = _buscar_columna(df, ["firstname", "first_name", "nombres"])
    col_ln = _buscar_columna(df, ["lastname", "last_name", "apellidos"])
    col_susp = _buscar_columna(df, ["suspended", "Suspended"])

    if not col_email and not col_user:
        raise ValueError("No se encontraron columnas email/username en la Nómina.")

    registros: list[dict[str, str]] = []
    by_email: dict[str, list[dict[str, str]]] = {}
    by_username: dict[str, list[dict[str, str]]] = {}
    by_rut: dict[str, list[dict[str, str]]] = {}

    for _, row in df.iterrows():
        email = normalizar_email(row.get(col_email)) if col_email else ""
        username = normalizar_email(row.get(col_user)) if col_user else ""
        firstname = str(row.get(col_fn) or "").strip() if col_fn else ""
        lastname = str(row.get(col_ln) or "").strip() if col_ln else ""
        nombre = f"{firstname} {lastname}".strip()
        rut = _rut_desde_fila(row, col_addr, col_id)
        suspended = str(row.get(col_susp) or "").strip() if col_susp else ""

        rec = {
            "username": username or str(row.get(col_user) or "").strip() if col_user else "",
            "email": email,
            "rut": rut,
            "nombre": nombre,
            "firstname": firstname,
            "lastname": lastname,
            "suspended": suspended,
        }
        registros.append(rec)

        if email:
            by_email.setdefault(email, []).append(rec)
        if username:
            by_username.setdefault(username, []).append(rec)
        if rut:
            by_rut.setdefault(rut, []).append(rec)
            # También indexar sin último dígito si parece incluir DV
            if len(rut) >= 8:
                by_rut.setdefault(rut[:-1], []).append(rec)

    return {
        "registros": registros,
        "by_email": by_email,
        "by_username": by_username,
        "by_rut": by_rut,
        "total": len(registros),
    }


def leer_asignacion(ruta: Path | str) -> pd.DataFrame:
    path = Path(ruta)
    if path.suffix.lower() not in (".xlsx", ".xls"):
        raise ValueError("El archivo de Asignación debe ser Excel (.xlsx/.xls).")

    xl = pd.ExcelFile(path)
    sheet = xl.sheet_names[0]
    for name in xl.sheet_names:
        if "asign" in name.casefold():
            sheet = name
            break

    df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    df.columns = [
        f"col_{i}" if (c is None or str(c).strip() == "" or str(c).startswith("Unnamed")) else str(c).strip()
        for i, c in enumerate(df.columns)
    ]

    col_email = _buscar_columna(df, ["EMAIL", "Email", "Correo"])
    col_rut = _buscar_columna(df, ["RUT", "Rut"])
    col_dig = _buscar_columna_digito(df)
    col_nombre = _buscar_columna(df, ["Nombre", "NOMBRE", "Nombres"])

    if not col_email and not col_rut:
        raise ValueError("No se encontraron columnas EMAIL o RUT en Asignación.")

    out = pd.DataFrame()
    mapping = {
        "Periodo": _buscar_columna(df, ["Periodo", "Período"]),
        "RUT": col_rut,
        "Digito": col_dig,
        "Nombre": col_nombre,
        "EMAIL": col_email,
        "Cargo": _buscar_columna(df, ["Cargo"]),
        "Unidad": _buscar_columna(df, ["Unidad"]),
        "Gcia/SubGcia": _buscar_columna(df, ["Gcia/SubGcia", "Gcia", "SubGcia"]),
        "VP": _buscar_columna(df, ["VP"]),
        "Ubicacion": _buscar_columna(df, ["Ubicación", "Ubicacion"]),
        "Nombre Jefe": _buscar_columna(df, ["Nombre Jefe", "Jefe"]),
        "TIPO CONTRATO": _buscar_columna(df, ["TIPO CONTRATO", "Tipo Contrato", "Contrato"]),
        "CeCo": _buscar_columna(df, ["CeCo", "CECO", "Centro Costo"]),
    }
    for dest, src in mapping.items():
        out[dest] = df[src].fillna("").astype(str).str.strip() if src else ""

    out["email_norm"] = out["EMAIL"].apply(normalizar_email)
    out["rut_norm"] = out["RUT"].apply(normalizar_rut)
    out["dig_norm"] = out["Digito"].apply(normalizar_rut)
    out["rut_dv"] = out.apply(
        lambda r: (r["rut_norm"] + r["dig_norm"]) if r["rut_norm"] and r["dig_norm"] else r["rut_norm"],
        axis=1,
    )

    out = out[(out["email_norm"] != "") | (out["rut_norm"] != "")].copy()
    return out


def _buscar_candidatos(row: pd.Series, indices: dict[str, Any]) -> tuple[list[dict[str, str]], str]:
    em = row["email_norm"]
    if em and em in indices["by_email"]:
        return indices["by_email"][em], "email"
    if em and em in indices["by_username"]:
        return indices["by_username"][em], "username"
    for key in (row["rut_dv"], row["rut_norm"]):
        if key and key in indices["by_rut"]:
            return indices["by_rut"][key], "rut"
    return [], ""


def _evaluar_identidad(
    row: pd.Series,
    candidatos: list[dict[str, str]],
    criterio: str,
) -> tuple[bool, list[dict[str, str]]]:
    """
    Devuelve (ok_identidad, inconsistencias).
    ok_identidad=True si al menos un candidato es coherente en RUT y nombre.
    """
    inconsistencias: list[dict[str, str]] = []
    hay_coherente = False

    # Email duplicado en nómina con distintas identidades
    if len(candidatos) > 1 and len({(c["rut"], normalizar_nombre(c["nombre"])) for c in candidatos}) > 1:
        for c in candidatos:
            inconsistencias.append(
                {
                    "Motivo": "Email/username duplicado en Nómina con distintas identidades",
                    "Criterio match": criterio,
                    "Asignacion EMAIL": row.get("EMAIL", ""),
                    "Asignacion RUT": f"{row.get('RUT', '')}-{row.get('Digito', '')}".strip("-"),
                    "Asignacion Nombre": row.get("Nombre", ""),
                    "Nomina username": c.get("username", ""),
                    "Nomina email": c.get("email", ""),
                    "Nomina RUT": c.get("rut", ""),
                    "Nomina Nombre": c.get("nombre", ""),
                    "Nomina suspended": c.get("suspended", ""),
                    "Detalle": f"Hay {len(candidatos)} registros en nómina para el mismo match",
                }
            )

    for c in candidatos:
        problemas: list[str] = []
        rut_cmp = ruts_coinciden(row["rut_norm"], row["rut_dv"], c.get("rut", ""))
        if rut_cmp is False:
            problemas.append(
                f"RUT distinto (Asignación {row['rut_dv'] or row['rut_norm']} vs Nómina {c.get('rut', '')})"
            )
        if not nombres_compatibles(str(row.get("Nombre") or ""), c.get("nombre", "")):
            problemas.append(
                f"Nombre distinto (Asignación '{row.get('Nombre', '')}' vs Nómina '{c.get('nombre', '')}')"
            )

        if not problemas:
            hay_coherente = True
        else:
            inconsistencias.append(
                {
                    "Motivo": "Datos no coinciden con la misma persona",
                    "Criterio match": criterio,
                    "Asignacion EMAIL": row.get("EMAIL", ""),
                    "Asignacion RUT": f"{row.get('RUT', '')}-{row.get('Digito', '')}".strip("-"),
                    "Asignacion Nombre": row.get("Nombre", ""),
                    "Nomina username": c.get("username", ""),
                    "Nomina email": c.get("email", ""),
                    "Nomina RUT": c.get("rut", ""),
                    "Nomina Nombre": c.get("nombre", ""),
                    "Nomina suspended": c.get("suspended", ""),
                    "Detalle": "; ".join(problemas),
                }
            )

    # Si todos los candidatos chocan, no hay identidad OK
    if candidatos and not hay_coherente:
        return False, inconsistencias

    # Si hay al menos uno coherente, filtramos inconsistencias de ese email duplicado
    # pero conservamos las de candidatos conflictivos + duplicados
    return True, inconsistencias


def _estilo_hoja(ws, fill_header: PatternFill):
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = font_body
            cell.border = border
    ws.row_dimensions[1].height = 26
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 4, 12), 45)
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _escribir_hoja(wb, titulo: str, columnas: list[str], filas: list[dict], fill: PatternFill):
    ws = wb.create_sheet(titulo)
    ws.append(columnas)
    for item in filas:
        ws.append([item.get(c, "") for c in columnas])
    _estilo_hoja(ws, fill)


def procesar_comparacion(ruta_asignacion: Path | str, ruta_nomina: Path | str, destino) -> dict:
    df_asig = leer_asignacion(ruta_asignacion)
    indices = leer_nomina(ruta_nomina)

    faltantes: list[dict] = []
    inconsistencias: list[dict] = []
    ok = 0

    for _, row in df_asig.iterrows():
        candidatos, criterio = _buscar_candidatos(row, indices)
        if not candidatos:
            faltantes.append({c: row.get(c, "") for c in COLS_FALTANTES})
            continue

        identidad_ok, issues = _evaluar_identidad(row, candidatos, criterio)
        inconsistencias.extend(issues)
        if identidad_ok:
            ok += 1
        else:
            # Email existe pero pertenece a otra persona → también reportar como faltante lógico
            faltantes.append({c: row.get(c, "") for c in COLS_FALTANTES})

    faltantes.sort(key=lambda x: (str(x.get("Nombre") or ""), str(x.get("EMAIL") or "")))
    inconsistencias.sort(
        key=lambda x: (str(x.get("Motivo") or ""), str(x.get("Asignacion Nombre") or ""))
    )

    wb = openpyxl.Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.append(["Métrica", "Valor"])
    ws_res.append(["Total en Asignación", len(df_asig)])
    ws_res.append(["Total en Nómina", indices["total"]])
    ws_res.append(["OK (misma persona)", ok])
    ws_res.append(["Faltan en Nómina", len(faltantes)])
    ws_res.append(["Inconsistencias (email/RUT/nombre)", len(inconsistencias)])
    fill_res = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    _estilo_hoja(ws_res, fill_res)

    fill_falt = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_inc = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    _escribir_hoja(wb, "Faltan en Nomina", COLS_FALTANTES, faltantes, fill_falt)
    _escribir_hoja(wb, "Inconsistencias", COLS_INCONSISTENCIAS, inconsistencias, fill_inc)

    wb.save(destino)
    return {
        "asignacion": len(df_asig),
        "nomina": indices["total"],
        "encontrados": ok,
        "faltan": len(faltantes),
        "inconsistencias": len(inconsistencias),
    }


def generar_reporte_bytes(
    ruta_asignacion: Path | str,
    ruta_nomina: Path | str,
) -> tuple[bytes, str, dict]:
    buf = io.BytesIO()
    stats = procesar_comparacion(ruta_asignacion, ruta_nomina, buf)
    filename = f"Reporte_Asignacion_Transelec_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf.getvalue(), filename, stats
