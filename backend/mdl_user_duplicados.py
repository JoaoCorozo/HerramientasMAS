"""Detección de usuarios duplicados en export mdl_user (Moodle)."""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DISPLAY_COLUMNS = [
    "id",
    "username",
    "firstname",
    "lastname",
    "email",
    "address",
    "idnumber",
    "institution",
    "department",
    "suspended",
    "deleted",
    "auth",
]

CRITERION_LABELS = {
    "address_idnumber": "Address / ID Number / Username",
    "email": "Correo",
}

CRITERION_SHEETS = {
    "address_idnumber": "Dup Addr IDN Username",
    "email": "Dup por Email",
}

COLOR_HEADER = "1F4E79"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_TITLE = "2E75B6"
COLOR_GROUP = "D6E4F0"
COLOR_ROW_ALT = "F2F7FB"
COLOR_BORDER = "B4C6DC"
COLOR_WARN = "FFC7CE"
COLOR_WARN_TEXT = "9C0006"
COLOR_MUTED = "E7E6E6"


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text in {"-", "NULL", "null", "None", "nan"}:
        return ""
    return text


def _normalize_key(value: Any) -> str:
    text = _normalize_value(value)
    if not text:
        return ""
    return re.sub(r"\s+", "", text).upper()


def _extract_column_keys(row: dict[str, Any], columns: tuple[str, ...]) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    for col in columns:
        if col == "email":
            email = _normalize_value(row.get("email")).lower()
            if email and "@" in email:
                keys.append(("email", email))
        else:
            value = _normalize_key(row.get(col))
            if value:
                keys.append((col, value))
    return keys


def _has_any_key(row: dict[str, Any]) -> bool:
    return bool(_extract_column_keys(row, ("address", "idnumber", "username", "email")))


def _no_key_reason(row: dict[str, Any]) -> str:
    address = _normalize_value(row.get("address"))
    idnumber = _normalize_value(row.get("idnumber"))
    username = _normalize_value(row.get("username"))
    email = _normalize_value(row.get("email"))

    parts: list[str] = []
    if not address or address in {"-", "NULL", "null"}:
        parts.append("address vacío o inválido")
    if not idnumber or idnumber in {"-", "NULL", "null"}:
        parts.append("idnumber vacío o inválido")
    if not username or username in {"-", "NULL", "null"}:
        parts.append("username vacío o inválido")
    if not email or "@" not in email:
        parts.append("email vacío o inválido")
    return "; ".join(parts) if parts else "Sin clave de comparación utilizable"


def read_mdl_user_file(path: str) -> tuple[list[str], list[dict[str, str]]]:
    ext = Path(path).suffix.lower()
    df = None

    if ext == ".csv":
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(path, sep=";", encoding=encoding, dtype=str, keep_default_na=False)
                break
            except Exception as exc:
                last_error = exc
        if df is None:
            raise ValueError(f"No se pudo leer el CSV: {last_error}")
    elif ext in {".xlsx", ".xls", ".xlsm"}:
        df = pd.read_excel(path, dtype=str, keep_default_na=False)
    else:
        raise ValueError("Formato no soportado. Use CSV (;) o Excel (.xlsx, .xls).")

    df = df.fillna("")
    headers = [str(col).strip() for col in df.columns]
    rows: list[dict[str, str]] = []
    for record in df.to_dict(orient="records"):
        clean = {str(k).strip(): _normalize_value(v) for k, v in record.items()}
        rows.append(clean)
    return headers, rows


class _UnionFind:
    def __init__(self, size: int) -> None:
        self.parent = list(range(size))

    def find(self, node: int) -> int:
        while self.parent[node] != node:
            self.parent[node] = self.parent[self.parent[node]]
            node = self.parent[node]
        return node

    def union(self, a: int, b: int) -> None:
        root_a, root_b = self.find(a), self.find(b)
        if root_a != root_b:
            self.parent[root_b] = root_a


def _find_duplicate_groups(
    rows: list[dict[str, str]],
    columns: tuple[str, ...],
    track: str,
) -> list[dict[str, Any]]:
    """Agrupa filas que comparten un mismo valor en las columnas indicadas."""
    n = len(rows)
    uf = _UnionFind(n)
    row_col_keys = [_extract_column_keys(row, columns) for row in rows]
    value_to_rows: dict[str, list[int]] = defaultdict(list)

    for idx, keys in enumerate(row_col_keys):
        for _col, val in keys:
            value_to_rows[val].append(idx)

    for indices in value_to_rows.values():
        unique_indices = sorted(set(indices))
        if len(unique_indices) < 2:
            continue
        anchor = unique_indices[0]
        for other in unique_indices[1:]:
            uf.union(anchor, other)

    components: dict[int, list[int]] = defaultdict(list)
    for idx, keys in enumerate(row_col_keys):
        if keys:
            components[uf.find(idx)].append(idx)

    groups: list[dict[str, Any]] = []
    allowed_cols = set(columns)

    for indices in components.values():
        if len(indices) < 2:
            continue

        group_rows = [rows[i] for i in sorted(indices)]
        linking_keys: list[str] = []
        columns_involved: set[str] = set()
        match_details: list[str] = []

        for val, all_indices in value_to_rows.items():
            in_group = sorted(set(all_indices) & set(indices))
            if len(in_group) < 2:
                continue
            linking_keys.append(val)
            cols_for_value: set[str] = set()
            for row_idx in in_group:
                for col, key_val in row_col_keys[row_idx]:
                    if key_val == val and col in allowed_cols:
                        cols_for_value.add(col)
                        columns_involved.add(col)
            if len(cols_for_value) > 1:
                match_details.append(f"{val}: {' <-> '.join(sorted(cols_for_value))}")
            elif cols_for_value:
                match_details.append(f"{val}: {next(iter(cols_for_value))}")

        linking_keys.sort()
        key_display = ", ".join(linking_keys[:2])
        if len(linking_keys) > 2:
            key_display += f" (+{len(linking_keys) - 2})"

        groups.append(
            {
                "track": track,
                "criterion": track,
                "criterion_label": CRITERION_LABELS.get(track, track),
                "key": key_display,
                "linking_keys": linking_keys,
                "columns_involved": sorted(columns_involved),
                "match_detail": "; ".join(match_details[:5]),
                "count": len(group_rows),
                "rows": group_rows,
            }
        )

    groups.sort(key=lambda g: (-g["count"], g["key"]))
    return groups


def analyze_duplicates(rows: list[dict[str, str]]) -> dict[str, Any]:
    if not rows:
        raise ValueError("El archivo no contiene filas de datos.")

    headers = list(rows[0].keys())
    address_col = next((h for h in headers if h.lower() == "address"), None)
    if not address_col:
        raise ValueError('No se encontró la columna "address" en el archivo.')

    rows_without_key_rows = [
        {**row, "_motivo_sin_clave": _no_key_reason(row)}
        for row in rows
        if not _has_any_key(row)
    ]

    groups_address_idnumber = _find_duplicate_groups(
        rows, ("address", "idnumber", "username"), "address_idnumber"
    )
    groups_email = _find_duplicate_groups(rows, ("email",), "email")

    duplicate_groups = groups_address_idnumber + groups_email
    duplicate_groups.sort(key=lambda g: (-g["count"], g["criterion"], g["key"]))

    groups_by_criterion = {
        "address_idnumber": groups_address_idnumber,
        "email": groups_email,
    }

    duplicate_rows = sum(group["count"] for group in duplicate_groups)
    sheet_group_counts = {
        track: len(groups_by_criterion[track]) for track in groups_by_criterion
    }
    sheet_row_counts = {
        track: sum(g["count"] for g in groups_by_criterion[track]) for track in groups_by_criterion
    }

    ui_columns = [col for col in DISPLAY_COLUMNS if col in headers]
    if not ui_columns:
        ui_columns = headers[:12]

    return {
        "total_rows": len(rows),
        "duplicate_groups": len(duplicate_groups),
        "duplicate_rows": duplicate_rows,
        "rows_without_key": len(rows_without_key_rows),
        "rows_without_key_rows": rows_without_key_rows,
        "by_criterion": {
            "address_idnumber": len(groups_address_idnumber),
            "email": len(groups_email),
        },
        "duplicate_rows_by_criterion": {
            "address_idnumber": sum(g["count"] for g in groups_address_idnumber),
            "email": sum(g["count"] for g in groups_email),
        },
        "sheet_group_counts": sheet_group_counts,
        "sheet_row_counts": sheet_row_counts,
        "groups": duplicate_groups,
        "groups_by_criterion": groups_by_criterion,
        "columns": ui_columns,
        "all_columns": headers,
    }


def build_result_payload(
    filename: str,
    analysis: dict[str, Any],
    scan_id: str | None = None,
) -> dict[str, Any]:
    return {
        "id": scan_id or datetime.now().strftime("%Y%m%d%H%M%S%f"),
        "filename": filename,
        "analyzed_at": datetime.now().isoformat(),
        **analysis,
    }


def _thin_border(color: str = COLOR_BORDER) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _style_header_row(ws, row_num: int, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(name="Calibri", bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = _fill(COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border("FFFFFF")


def _autofit_columns(ws, num_cols: int, max_row: int, min_width: int = 10, max_width: int = 50) -> None:
    for col_idx in range(1, num_cols + 1):
        col = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col].width = min(max(max_len + 2, min_width), max_width)


def _write_sheet_title(ws, title: str, subtitle: str, num_cols: int) -> int:
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=13, bold=True, color=COLOR_HEADER_TEXT)
    title_cell.fill = _fill(COLOR_TITLE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = subtitle
    subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color="1F2937")
    subtitle_cell.fill = _fill(COLOR_MUTED)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6
    return 4


def _write_duplicate_groups_sheet(
    ws,
    sheet_title: str,
    subtitle: str,
    groups: list[dict[str, Any]],
    data_columns: list[str],
) -> None:
    meta_columns = [
        "Grupo",
        "Clave(s) duplicada(s)",
        "Detalle cruce columnas",
        "Columnas involucradas",
        "Registros en grupo",
    ]
    columns = meta_columns + data_columns
    header_row = _write_sheet_title(ws, sheet_title, subtitle, len(columns))

    ws.append(columns)
    _style_header_row(ws, header_row, len(columns))

    if not groups:
        ws.append(["—", "—", "—", "—", 0, *([""] * len(data_columns))])
        note_row = header_row + 2
        ws.merge_cells(f"A{note_row}:{get_column_letter(len(columns))}{note_row}")
        note_cell = ws.cell(row=note_row, column=1)
        note_cell.value = "No se encontraron duplicados para este criterio."
        note_cell.font = Font(italic=True, color="666666")
        _autofit_columns(ws, len(columns), note_row)
        ws.sheet_view.showGridLines = False
        return

    row_idx = header_row + 1
    for group_num, group in enumerate(groups, start=1):
        group_label = f"Grupo {group_num}"
        for member_idx, row in enumerate(group["rows"]):
            values = [
                group_label if member_idx == 0 else "",
                group.get("key", ""),
                group.get("match_detail", "") if member_idx == 0 else "",
                ", ".join(group.get("columns_involved") or []) if member_idx == 0 else "",
                group.get("count", len(group["rows"])) if member_idx == 0 else "",
            ]
            values.extend(row.get(col, "") for col in data_columns)
            ws.append(values)

            fill_color = COLOR_GROUP if member_idx == 0 else (COLOR_ROW_ALT if row_idx % 2 else "FFFFFF")
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = _thin_border()
                cell.fill = _fill(fill_color)
                cell.alignment = Alignment(vertical="center", wrap_text=col_idx > 3)
            row_idx += 1

    ws.freeze_panes = f"A{header_row + 1}"
    if row_idx > header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{row_idx - 1}"
    _autofit_columns(ws, len(columns), row_idx - 1)
    ws.sheet_view.showGridLines = False


def _write_no_key_sheet(ws, rows: list[dict[str, str]], data_columns: list[str]) -> None:
    columns = ["Motivo sin clave"] + data_columns
    header_row = _write_sheet_title(
        ws,
        "Filas sin clave utilizable",
        "Usuarios que no tienen address, idnumber ni email válidos para comparar duplicados.",
        len(columns),
    )

    ws.append(columns)
    _style_header_row(ws, header_row, len(columns))

    row_idx = header_row + 1
    if not rows:
        ws.append(["—", *([""] * len(data_columns))])
        note_row = header_row + 2
        ws.merge_cells(f"A{note_row}:{get_column_letter(len(columns))}{note_row}")
        note_cell = ws.cell(row=note_row, column=1)
        note_cell.value = "Todas las filas tenían al menos una clave utilizable."
        note_cell.font = Font(italic=True, color="666666")
        _autofit_columns(ws, len(columns), note_row)
        ws.sheet_view.showGridLines = False
        return

    for offset, row in enumerate(rows):
        values = [row.get("_motivo_sin_clave", _no_key_reason(row))]
        values.extend(row.get(col, "") for col in data_columns)
        ws.append(values)
        fill_color = COLOR_WARN if offset % 2 == 0 else "FFFFFF"
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _thin_border()
            cell.fill = _fill(fill_color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 1:
                cell.font = Font(color=COLOR_WARN_TEXT)
        row_idx += 1

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{row_idx - 1}"
    _autofit_columns(ws, len(columns), row_idx - 1)
    ws.sheet_view.showGridLines = False


def _write_summary_sheet(ws, analysis: dict[str, Any]) -> None:
    ws.title = "Resumen"
    by_crit = analysis.get("by_criterion", {})
    rows_by_crit = analysis.get("duplicate_rows_by_criterion", {})

    summary_rows = [
        ("Total filas analizadas", analysis.get("total_rows", 0)),
        ("Total grupos duplicados", analysis.get("duplicate_groups", 0)),
        ("Total filas en duplicados", analysis.get("duplicate_rows", 0)),
        ("Filas sin clave utilizable", analysis.get("rows_without_key", 0)),
        ("", ""),
        ("Grupos address/idnumber/username", by_crit.get("address_idnumber", 0)),
        ("Filas en duplicados address/idnumber/username", rows_by_crit.get("address_idnumber", 0)),
        ("Grupos duplicados por email", by_crit.get("email", 0)),
        ("Filas en duplicados por email", rows_by_crit.get("email", 0)),
        ("", ""),
        ("Regla principal", "Cruce entre address, idnumber y username"),
        ("Regla email", "Duplicados en columna email"),
        ("", ""),
        ("Archivo origen", analysis.get("filename", "")),
        ("Fecha de análisis", analysis.get("analyzed_at", "")),
        ("", ""),
        ("Hojas del reporte", "Resumen"),
        ("", "Dup Addr IDN Username"),
        ("", "Dup por Email"),
        ("", "Sin clave utilizable"),
        ("", "Duplicados (todos)"),
    ]

    ws["A1"] = "Reporte de usuarios duplicados — mdl_user"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_TEXT)
    ws["A1"].fill = _fill(COLOR_TITLE)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    start_row = 3
    ws.cell(row=start_row, column=1, value="Indicador").font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="Valor").font = Font(bold=True)
    for offset, (label, value) in enumerate(summary_rows, start=1):
        ws.cell(row=start_row + offset, column=1, value=label)
        ws.cell(row=start_row + offset, column=2, value=value)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 52
    ws.sheet_view.showGridLines = False


def build_export_filename(source_filename: str | None = None) -> str:
    stem = Path(source_filename or "mdl_user").stem
    safe_stem = re.sub(r"[^\w.-]+", "_", stem)[:40] or "mdl_user"
    return f"Duplicados_{safe_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def build_excel_bytes(analysis: dict[str, Any]) -> tuple[bytes, str]:
    wb = openpyxl.Workbook()
    data_columns = list(analysis.get("all_columns") or analysis.get("columns") or DISPLAY_COLUMNS)

    groups_by_criterion = analysis.get("groups_by_criterion")
    if not groups_by_criterion:
        groups_by_criterion = {"address_idnumber": [], "email": []}
        for group in analysis.get("groups") or []:
            track = group.get("track") or group.get("criterion")
            if track in groups_by_criterion:
                groups_by_criterion[track].append(group)
            elif group.get("criterion") in ("address", "idnumber"):
                groups_by_criterion["address_idnumber"].append(group)
            elif group.get("criterion") == "email":
                groups_by_criterion["email"].append(group)

    _write_summary_sheet(wb.active, analysis)

    criterion_subtitles = {
        "address_idnumber": (
            "Usuarios agrupados cuando un mismo valor aparece en address, idnumber y/o username "
            "(cruce entre las tres columnas)."
        ),
        "email": "Usuarios agrupados cuando comparten el mismo correo en la columna email.",
    }

    for criterion, sheet_name in CRITERION_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        groups = groups_by_criterion.get(criterion, [])
        _write_duplicate_groups_sheet(
            ws,
            f"Duplicados — {CRITERION_LABELS[criterion]}",
            f"{criterion_subtitles[criterion]} "
            f"{analysis.get('sheet_group_counts', {}).get(criterion, len(groups))} grupo(s) en esta hoja.",
            groups,
            data_columns,
        )

    ws_no_key = wb.create_sheet("Sin clave utilizable")
    no_key_rows = analysis.get("rows_without_key_rows") or []
    _write_no_key_sheet(ws_no_key, no_key_rows, data_columns)

    ws_all = wb.create_sheet("Duplicados (todos)")
    _write_duplicate_groups_sheet(
        ws_all,
        "Todos los duplicados",
        f"Listado consolidado de los {analysis.get('duplicate_groups', 0)} grupos detectados.",
        analysis.get("groups") or [],
        data_columns,
    )

    filename = build_export_filename(analysis.get("filename"))
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename
