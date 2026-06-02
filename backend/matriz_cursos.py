"""
Matriz «cursos bex Moodle»:
  - Hoja catálogo: id + shortname (+ fullname)
  - Una hoja por perfil de inducción: columna id (encabezado «id» o solo números)
"""
from __future__ import annotations

import os
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import HTTPException

from paths import resolve_all_matriz_paths, resolve_matriz_cursos_path

try:
    import xlrd
except ImportError:
    xlrd = None  # type: ignore

PERFIL_COLUMN_HINTS = (
    "perfil de induccion",
    "perfil de inducci",
    "perfil para induccion",
    "perfil induccion",
    "perfil de inducción",
    "perfil",
    "department",
    "departamento",
)

_cache: dict[str, Any] = {"mtime": None, "path": None, "catalog": {}, "profiles": {}}


def normalize_text(val: str | None) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )
    return s.upper()


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _normalize_id(val: Any) -> str | None:
    s = _cell_str(val)
    if not s:
        return None
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return None


def _header_map(header_row: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _cell_str(cell).lower().replace(" ", "")
        if key:
            mapping[key] = idx
    return mapping


def _is_catalog_sheet(sheet_name: str, headers: dict[str, int]) -> bool:
    name = sheet_name.lower()
    if "mdl_course" in name or "catalogo" in name or name in ("catálogo", "courses"):
        return True
    return "shortname" in headers and "id" in headers


def _read_xls_sheets(path: str) -> dict[str, list[tuple[Any, ...]]]:
    if xlrd is None:
        raise HTTPException(status_code=500, detail="Falta xlrd para leer .xls")
    wb = xlrd.open_workbook(path)
    out: dict[str, list[tuple[Any, ...]]] = {}
    for sname in wb.sheet_names():
        sh = wb.sheet_by_name(sname)
        rows = [
            tuple(sh.cell_value(r, c) for c in range(sh.ncols))
            for r in range(sh.nrows)
        ]
        out[sname] = rows
    return out


def _read_xlsx_sheets(path: str) -> dict[str, list[tuple[Any, ...]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[tuple[Any, ...]]] = {}
    try:
        for sname in wb.sheetnames:
            sh = wb[sname]
            out[sname] = [tuple(row) for row in sh.iter_rows(values_only=True)]
    finally:
        wb.close()
    return out


def read_workbook_sheets(path: str) -> dict[str, list[tuple[Any, ...]]]:
    if Path(path).suffix.lower() == ".xls":
        return _read_xls_sheets(path)
    return _read_xlsx_sheets(path)


def _read_ids_from_profile_sheet(rows: list[tuple[Any, ...]]) -> list[str]:
    """Lee ids de curso desde una hoja de perfil (con o sin fila de encabezado)."""
    if not rows:
        return []

    headers = _header_map(rows[0])
    ids: list[str] = []

    if "id" in headers:
        id_idx = headers["id"]
        for row in rows[1:]:
            if id_idx < len(row):
                cid = _normalize_id(row[id_idx])
                if cid:
                    ids.append(cid)
        return ids

    # Sin encabezado: primera columna con ids numéricos
    first_id = _normalize_id(rows[0][0]) if rows[0] else None
    start = 0 if first_id else 1
    for row in rows[start:]:
        if not row:
            continue
        cid = _normalize_id(row[0])
        if cid:
            ids.append(cid)
    return ids


def _build_catalog(sheets: dict[str, list[tuple[Any, ...]]]) -> dict[str, str]:
    catalog: dict[str, str] = {}
    for sname, rows in sheets.items():
        if not rows:
            continue
        headers = _header_map(rows[0])
        if not _is_catalog_sheet(sname, headers):
            continue
        id_idx = headers.get("id")
        sn_idx = headers.get("shortname")
        if id_idx is None or sn_idx is None:
            continue
        for row in rows[1:]:
            if id_idx >= len(row):
                continue
            cid = _normalize_id(row[id_idx])
            if not cid:
                continue
            shortname = _cell_str(row[sn_idx]) if sn_idx < len(row) else ""
            if shortname:
                catalog[cid] = shortname
    return catalog


def _build_profile_index(
    sheets: dict[str, list[tuple[Any, ...]]],
) -> dict[str, dict[str, Any]]:
    """Clave normalizada del nombre de hoja -> {sheet_name, ids}."""
    index: dict[str, dict[str, Any]] = {}
    for sname, rows in sheets.items():
        if not rows:
            continue
        headers = _header_map(rows[0])
        if _is_catalog_sheet(sname, headers):
            continue
        ids = _read_ids_from_profile_sheet(rows)
        if not ids:
            continue
        key = normalize_text(sname)
        index[key] = {"sheet_name": sname.strip(), "ids": ids}
    return index


def _load_cached() -> tuple[dict[str, str], dict[str, dict[str, Any]], str]:
    paths = resolve_all_matriz_paths()
    if not paths:
        return {}, {}, ""

    try:
        mtime = max(os.path.getmtime(p) for p in paths)
    except OSError:
        mtime = None

    cache_key = "|".join(paths)
    if (
        _cache.get("mtime") == mtime
        and _cache.get("path") == cache_key
        and _cache.get("catalog")
        and _cache.get("profiles")
    ):
        return _cache["catalog"], _cache["profiles"], cache_key

    catalog: dict[str, str] = {}
    profiles: dict[str, dict[str, Any]] = {}
    for path in paths:
        sheets = read_workbook_sheets(path)
        catalog.update(_build_catalog(sheets))
        profiles.update(_build_profile_index(sheets))

    _cache["mtime"] = mtime
    _cache["path"] = cache_key
    _cache["catalog"] = catalog
    _cache["profiles"] = profiles
    return catalog, profiles, paths[0]


def invalidate_cache() -> None:
    _cache["mtime"] = None


def get_matriz_info() -> dict[str, Any]:
    path = resolve_matriz_cursos_path()
    if not path:
        return {
            "loaded": False,
            "path": None,
            "error": "No se encontró «cursos bex Moodle.xlsx» o «.xls»",
        }

    catalog, profiles, _ = _load_cached()
    perfiles = []
    for key, data in sorted(profiles.items()):
        ids = data["ids"]
        shortnames = [catalog.get(i, f"(id {i} sin shortname)") for i in ids]
        perfiles.append(
            {
                "hoja": data["sheet_name"],
                "clave": key,
                "cantidad_ids": len(ids),
                "cursos": shortnames[:5],
            }
        )

    return {
        "loaded": True,
        "path": path,
        "archivos": resolve_all_matriz_paths(),
        "catalogo_cursos": len(catalog),
        "perfiles": perfiles,
    }


def courses_for_perfil(perfil_norm: str, perfil_display: str = "") -> list[str]:
    catalog, profiles, _ = _load_cached()
    if not catalog or not profiles:
        return []

    profile_data = profiles.get(perfil_norm)
    if not profile_data and perfil_display:
        profile_data = profiles.get(normalize_text(perfil_display))

    if not profile_data:
        for key, data in profiles.items():
            if perfil_norm == key or perfil_norm in key or key in perfil_norm:
                profile_data = data
                break
            if perfil_display and normalize_text(perfil_display) == key:
                profile_data = data
                break

    if not profile_data:
        return []

    shortnames: list[str] = []
    for cid in profile_data["ids"]:
        sn = catalog.get(cid)
        if sn:
            shortnames.append(sn)
    return shortnames


def load_mapa_cursos_perfil(require_file: bool = False) -> dict[str, list[str]]:
    """Compatibilidad: perfil normalizado -> lista de shortnames."""
    path = resolve_matriz_cursos_path()
    if not path:
        if require_file:
            raise HTTPException(
                status_code=404,
                detail='No se encontró «cursos bex Moodle» (.xlsx / .xls).',
            )
        return {}

    catalog, profiles, _ = _load_cached()
    if require_file and not catalog:
        raise HTTPException(
            status_code=500,
            detail="Falta hoja catálogo con columnas id y shortname.",
        )
    if require_file and not profiles:
        raise HTTPException(
            status_code=404,
            detail="No hay hojas de perfil con columna id.",
        )

    mapa: dict[str, list[str]] = {}
    for key, data in profiles.items():
        mapa[key] = courses_for_perfil(key, data["sheet_name"])
    return mapa


def _normalize_col_name(name: str) -> str:
    return normalize_text(name).replace(" ", "")


def extract_perfil_from_row(
    row: tuple[Any, ...] | list[Any],
    col_name_to_idx: dict[str, int],
    processed_department: str = "",
) -> tuple[str, str]:
    if processed_department and str(processed_department).strip():
        original = str(processed_department).strip()
        return original, normalize_text(original)

    for hint in PERFIL_COLUMN_HINTS:
        hint_key = _normalize_col_name(hint)
        for col_name, idx in col_name_to_idx.items():
            col_key = _normalize_col_name(col_name)
            if hint_key in col_key or col_key in hint_key:
                if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
                    original = str(row[idx]).strip()
                    return original, normalize_text(original)

    return "", "SIN_PERFIL"


def resolve_courses_for_perfil(
    mapa: dict[str, list[str]],
    perfil_norm: str,
    perfil_display: str = "",
) -> list[str]:
    """Usa la matriz en caché; mapa se ignora si hay datos frescos."""
    return courses_for_perfil(perfil_norm, perfil_display)
