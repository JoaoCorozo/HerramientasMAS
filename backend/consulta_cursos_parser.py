"""Parsea texto pegado desde consultas Moodle y genera Excel."""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta reporte conectividad
COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_TITLE_BG = "2E75B6"
COLOR_SUBTITLE_TEXT = "D6DCE4"
COLOR_ROW_ALT = "F2F7FB"
COLOR_ROW_WHITE = "FFFFFF"
COLOR_BORDER = "B4C6DC"
COLOR_ESTADO_OK = "C6EFCE"
COLOR_ESTADO_OK_TEXT = "006100"
COLOR_ESTADO_WARN = "FFEB9C"
COLOR_ESTADO_WARN_TEXT = "9C6500"
COLOR_ESTADO_BAD = "FFC7CE"
COLOR_ESTADO_BAD_TEXT = "9C0006"
COLOR_INSCRITO_SI = "E2F0D9"
COLOR_INSCRITO_NO = "FCE4D6"
COLOR_AVANCE_FULL = "C6EFCE"
COLOR_AVANCE_ZERO = "FFC7CE"
COLOR_AVANCE_PARTIAL = "FFEB9C"

COL_ESTADO = 8
COL_INSCRITO = 6
COL_AVANCE = 9
COL_CUENTA = 5
HEADER_ROW = 4
DATA_START_ROW = 5

COURSE_RE = re.compile(r"^Curso:\s*(.+?)\s*\(ID\s*(\d+)\)\s*$", re.IGNORECASE)
HEADER_PREFIX = "Usuario\tNombre\tApellido"
USER_ID_RE = re.compile(r"^\d{7,10}$")

DATA_COLUMNS = [
    "Usuario",
    "Nombre",
    "Apellido",
    "Correo",
    "Cuenta",
    "Inscrito",
    "Nota",
    "Estado",
    "Avance",
    "Fecha Inscripción/Baja",
    "Responsable",
]

@dataclass
class ConsultaRow:
    curso_id: str
    curso: str
    usuario: str
    nombre: str
    apellido: str
    correo: str
    cuenta: str
    inscrito: str
    nota: str
    estado: str
    avance: str
    fecha_inscripcion: str
    responsable: str

    def as_data_list(self) -> list[str]:
        return [
            self.usuario,
            self.nombre,
            self.apellido,
            self.correo,
            self.cuenta,
            self.inscrito,
            self.nota,
            self.estado,
            self.avance,
            self.fecha_inscripcion,
            self.responsable,
        ]


def _should_skip_line(line: str) -> bool:
    if not line:
        return True
    lower = line.lower()
    if line in {"Anterior", "Siguiente", "Buscar:"}:
        return True
    if lower.startswith("resultado consulta"):
        return True
    if lower.startswith("mostrando "):
        return True
    if re.fullmatch(r"\d{1,3}", line):
        return True
    if line.startswith(HEADER_PREFIX):
        return True
    return False


def _parse_data_line(line: str, curso: str, curso_id: str) -> ConsultaRow | None:
    parts = line.split("\t")
    if not parts or not USER_ID_RE.match(parts[0].strip()):
        return None
    while len(parts) < len(DATA_COLUMNS):
        parts.append("")
    parts = parts[: len(DATA_COLUMNS)]
    return ConsultaRow(
        curso_id=curso_id,
        curso=curso,
        usuario=parts[0].strip(),
        nombre=parts[1].strip(),
        apellido=parts[2].strip(),
        correo=parts[3].strip(),
        cuenta=parts[4].strip(),
        inscrito=parts[5].strip(),
        nota=parts[6].strip(),
        estado=parts[7].strip(),
        avance=parts[8].strip(),
        fecha_inscripcion=parts[9].strip(),
        responsable=parts[10].strip(),
    )


def parse_consulta_text(text: str) -> list[ConsultaRow]:
    rows: list[ConsultaRow] = []
    curso = ""
    curso_id = ""

    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    for raw_line in normalized.split("\n"):
        line = raw_line.strip()
        if _should_skip_line(line):
            continue

        course_match = COURSE_RE.match(line)
        if course_match:
            curso = course_match.group(1).strip()
            curso_id = course_match.group(2).strip()
            continue

        if not curso_id:
            continue

        parsed = _parse_data_line(line, curso, curso_id)
        if parsed:
            rows.append(parsed)

    return rows


def summarize_rows(rows: list[ConsultaRow]) -> dict:
    cursos: dict[str, dict] = {}
    usuarios: set[str] = set()
    for row in rows:
        usuarios.add(row.usuario)
        key = row.curso_id
        if key not in cursos:
            cursos[key] = {"curso_id": row.curso_id, "curso": row.curso, "registros": 0}
        cursos[key]["registros"] += 1

    return {
        "total_registros": len(rows),
        "total_cursos": len(cursos),
        "total_usuarios": len(usuarios),
        "cursos": sorted(cursos.values(), key=lambda c: c["curso"]),
    }


_SHEET_INVALID_CHARS = re.compile(r"[\[\]:\\/?*]")


def _unique_sheet_title(curso_id: str, curso: str, used: set[str]) -> str:
    base = _SHEET_INVALID_CHARS.sub("", f"{curso_id} - {curso}").strip()
    if not base:
        base = f"Curso {curso_id}"
    if len(base) > 31:
        base = base[:31].rstrip()

    title = base
    suffix = 2
    while title in used:
        extra = f" ({suffix})"
        title = base[: 31 - len(extra)] + extra
        suffix += 1
    used.add(title)
    return title


def _thin_border(color: str = COLOR_BORDER) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _normalize_status(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def _estado_style(estado: str) -> tuple[PatternFill, Font]:
    norm = _normalize_status(estado)
    if "aprobado" in norm:
        return _fill(COLOR_ESTADO_OK), Font(color=COLOR_ESTADO_OK_TEXT, bold=True)
    if "sin actividad" in norm or norm == "n/a":
        return _fill(COLOR_ESTADO_BAD), Font(color=COLOR_ESTADO_BAD_TEXT, bold=True)
    if "pendiente" in norm or "progreso" in norm:
        return _fill(COLOR_ESTADO_WARN), Font(color=COLOR_ESTADO_WARN_TEXT, bold=True)
    return _fill(COLOR_ROW_WHITE), Font(color="333333")


def _avance_style(avance: str) -> PatternFill:
    text = (avance or "").strip().replace("%", "")
    try:
        pct = float(text)
    except ValueError:
        return _fill(COLOR_ROW_WHITE)
    if pct >= 100:
        return _fill(COLOR_AVANCE_FULL)
    if pct <= 0:
        return _fill(COLOR_AVANCE_ZERO)
    return _fill(COLOR_AVANCE_PARTIAL)


def _yes_no_style(value: str, positive: str = "si") -> PatternFill:
    norm = _normalize_status(value)
    if norm == positive:
        return _fill(COLOR_INSCRITO_SI)
    if norm in {"no", "n/a"}:
        return _fill(COLOR_INSCRITO_NO)
    return _fill(COLOR_ROW_WHITE)


def _group_rows_by_course(rows: list[ConsultaRow]) -> list[tuple[tuple[str, str], list[ConsultaRow]]]:
    groups: list[tuple[tuple[str, str], list[ConsultaRow]]] = []
    current_key: tuple[str, str] | None = None
    current_rows: list[ConsultaRow] = []

    for row in rows:
        key = (row.curso_id, row.curso)
        if key != current_key:
            if current_rows and current_key is not None:
                groups.append((current_key, current_rows))
            current_key = key
            current_rows = [row]
        else:
            current_rows.append(row)

    if current_rows and current_key is not None:
        groups.append((current_key, current_rows))
    return groups


def _populate_styled_sheet(
    ws,
    curso_id: str,
    curso: str,
    course_rows: list[ConsultaRow],
    generated_at: datetime,
) -> None:
    num_cols = len(DATA_COLUMNS)
    last_col = get_column_letter(num_cols)

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = f"Reporte de Conectividad — {curso}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_TEXT)
    title_cell.fill = _fill(COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = (
        f"ID Curso: {curso_id}  |  "
        f"Usuarios: {len(course_rows)}  |  "
        f"Generado: {generated_at.strftime('%d/%m/%Y %H:%M')}"
    )
    subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color="1F2937")
    subtitle_cell.fill = _fill(COLOR_SUBTITLE_TEXT)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6

    for col_idx, header in enumerate(DATA_COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = _fill(COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border("FFFFFF")

    for row_offset, data_row in enumerate(course_rows):
        excel_row = DATA_START_ROW + row_offset
        values = data_row.as_data_list()
        row_fill = _fill(COLOR_ROW_ALT if row_offset % 2 else COLOR_ROW_WHITE)

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10, color="1F2937")
            cell.fill = row_fill
            cell.border = _thin_border()
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=col_idx in {4, 10},
            )

        estado_cell = ws.cell(row=excel_row, column=COL_ESTADO)
        estado_fill, estado_font = _estado_style(str(estado_cell.value or ""))
        estado_cell.fill = estado_fill
        estado_cell.font = estado_font
        estado_cell.alignment = Alignment(horizontal="center", vertical="center")

        inscrito_cell = ws.cell(row=excel_row, column=COL_INSCRITO)
        inscrito_cell.fill = _yes_no_style(str(inscrito_cell.value or ""))
        inscrito_cell.alignment = Alignment(horizontal="center", vertical="center")
        inscrito_cell.font = Font(name="Calibri", size=10, bold=True, color="1F2937")

        cuenta_cell = ws.cell(row=excel_row, column=COL_CUENTA)
        if _normalize_status(str(cuenta_cell.value or "")) == "activa":
            cuenta_cell.fill = _fill(COLOR_INSCRITO_SI)
        cuenta_cell.alignment = Alignment(horizontal="center", vertical="center")

        avance_cell = ws.cell(row=excel_row, column=COL_AVANCE)
        avance_cell.fill = _avance_style(str(avance_cell.value or ""))
        avance_cell.alignment = Alignment(horizontal="center", vertical="center")
        avance_cell.font = Font(name="Calibri", size=10, bold=True, color="1F2937")

        ws.cell(row=excel_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(DATA_COLUMNS, start=1):
        col = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(HEADER_ROW, DATA_START_ROW + len(course_rows)):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col].width = min(max(max_len + 2, 12), 55)

    ws.freeze_panes = f"A{DATA_START_ROW}"
    if course_rows:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{DATA_START_ROW + len(course_rows) - 1}"

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"${HEADER_ROW}:${HEADER_ROW}"


def build_excel_bytes(rows: list[ConsultaRow]) -> tuple[bytes, str]:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    generated_at = datetime.now()

    for (curso_id, curso), course_rows in _group_rows_by_course(rows):
        title = _unique_sheet_title(curso_id, curso, used_titles)
        ws = wb.create_sheet(title=title)
        _populate_styled_sheet(ws, curso_id, curso, course_rows, generated_at)

    filename = f"Consulta_Cursos_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename
