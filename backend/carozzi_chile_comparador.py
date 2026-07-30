"""Comparador Todos Chile (Carozzi): reporte plataforma vs export BBDD por RUT/username."""

from __future__ import annotations

import csv
import io
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import BinaryIO

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COLS_PLATAFORMA = ["RUT", "NOMBRES", "APELLIDOS", "EMAIL", "NOMBRE DE CARGO", "EMAIL CAROZZI"]
COLS_BBDD = ["id", "username", "firstname", "lastname", "email", "email_carozzi", "email_personal"]


def normalizar_rut(valor) -> str:
    """RUT/username comparable: solo dígitos y K, mayúsculas, sin puntos ni guión."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    s = str(valor).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace("\xa0", "").replace(" ", "")
    s = re.sub(r"[^0-9Kk]", "", s).upper()
    return s


def _detectar_encoding(ruta: Path | str) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(ruta, encoding=encoding) as f:
                f.readline()
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detectar_separador(texto: str) -> str:
    primera = texto.splitlines()[0] if texto else ""
    try:
        dialecto = csv.Sniffer().sniff(texto[:8192], delimiters=";,")
        if dialecto.delimiter in (";", ","):
            return dialecto.delimiter
    except csv.Error:
        pass
    return ";" if primera.count(";") >= primera.count(",") else ","


def _leer_csv(ruta: Path | str) -> pd.DataFrame:
    encoding = _detectar_encoding(ruta)
    with open(ruta, encoding=encoding, errors="replace") as f:
        texto = f.read()
    sep = _detectar_separador(texto)
    kwargs = dict(sep=sep, dtype=str, engine="python", quotechar='"', skipinitialspace=True)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=pd.errors.ParserWarning)
        try:
            return pd.read_csv(io.StringIO(texto), on_bad_lines="warn", **kwargs)
        except TypeError:
            return pd.read_csv(
                io.StringIO(texto),
                error_bad_lines=False,
                warn_bad_lines=False,
                **kwargs,
            )


def _buscar_columna(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    mapa = {str(c).strip().casefold(): c for c in df.columns}
    for cand in candidatos:
        key = cand.strip().casefold()
        if key in mapa:
            return mapa[key]
    for col in df.columns:
        limpio = re.sub(r"[^a-z0-9]", "", str(col).casefold())
        for cand in candidatos:
            cand_limpo = re.sub(r"[^a-z0-9]", "", cand.casefold())
            if limpio == cand_limpo:
                return col
    return None


def leer_plataforma_todos(ruta: Path | str) -> pd.DataFrame:
    path = Path(ruta)
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(path, dtype=str)
    elif ext == ".csv":
        df = _leer_csv(path)
    else:
        raise ValueError("El archivo Todos (plataforma) debe ser Excel (.xlsx/.xls) o CSV.")

    col_rut = _buscar_columna(df, ["RUT", "Rut", "rut"])
    if not col_rut:
        raise ValueError("No se encontró la columna RUT en el archivo Todos (plataforma).")

    col_map = {
        "RUT": col_rut,
        "NOMBRES": _buscar_columna(df, ["NOMBRES", "Nombres", "Nombre"]),
        "APELLIDOS": _buscar_columna(df, ["APELLIDOS", "Apellidos", "Apellido"]),
        "EMAIL": _buscar_columna(df, ["EMAIL", "Email", "Correo"]),
        "NOMBRE DE CARGO": _buscar_columna(df, ["NOMBRE DE CARGO", "Nombre de Cargo", "Cargo"]),
        "EMAIL CAROZZI": _buscar_columna(df, ["EMAIL CAROZZI", "Email Carozzi", "Correo Carozzi"]),
    }

    out = pd.DataFrame()
    for dest, src in col_map.items():
        out[dest] = df[src].fillna("").astype(str).str.strip() if src else ""
    out["RUT_norm"] = out["RUT"].apply(normalizar_rut)
    out = out[out["RUT_norm"] != ""].copy()
    out = out.drop_duplicates(subset=["RUT_norm"], keep="first")
    return out


def leer_bbdd_ultimos(ruta: Path | str) -> pd.DataFrame:
    path = Path(ruta)
    if path.suffix.lower() != ".csv":
        raise ValueError("El archivo de BBDD (últimos) debe ser CSV (.csv).")

    df = _leer_csv(path)
    col_user = _buscar_columna(df, ["username", "Username", "RUT", "rut"])
    if not col_user:
        raise ValueError("No se encontró la columna username en el CSV de BBDD.")

    col_map = {
        "id": _buscar_columna(df, ["id", "ID"]),
        "username": col_user,
        "firstname": _buscar_columna(df, ["firstname", "first_name", "Nombres"]),
        "lastname": _buscar_columna(df, ["lastname", "last_name", "Apellidos"]),
        "email": _buscar_columna(df, ["email", "Email"]),
        "email_carozzi": _buscar_columna(df, ["email_carozzi", "Email Carozzi"]),
        "email_personal": _buscar_columna(df, ["email_personal", "Email Personal"]),
    }

    out = pd.DataFrame()
    for dest, src in col_map.items():
        out[dest] = df[src].fillna("").astype(str).str.strip() if src else ""
    out["RUT_norm"] = out["username"].apply(normalizar_rut)
    out = out[out["RUT_norm"] != ""].copy()
    out = out.drop_duplicates(subset=["RUT_norm"], keep="first")
    return out


def _estilo_hoja(ws, fill_header: PatternFill):
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = font_body
            cell.border = border
    ws.row_dimensions[1].height = 26
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _escribir_hoja(wb, titulo: str, columnas: list[str], filas: list[dict], fill: PatternFill):
    ws = wb.create_sheet(titulo)
    ws.append(columnas)
    for item in filas:
        ws.append([item.get(c, "") for c in columnas])
    _estilo_hoja(ws, fill)


def procesar_comparacion(
    ruta_plataforma: Path | str,
    ruta_bbdd: Path | str,
    destino: Path | str | BinaryIO,
) -> dict:
    df_plat = leer_plataforma_todos(ruta_plataforma)
    df_bbdd = leer_bbdd_ultimos(ruta_bbdd)

    set_plat = set(df_plat["RUT_norm"])
    set_bbdd = set(df_bbdd["RUT_norm"])

    comunes = set_plat & set_bbdd
    solo_plat = set_plat - set_bbdd
    solo_bbdd = set_bbdd - set_plat

    plat_idx = df_plat.set_index("RUT_norm")
    bbdd_idx = df_bbdd.set_index("RUT_norm")

    coincidencias = []
    for rut in sorted(comunes):
        row = plat_idx.loc[rut]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        coincidencias.append({c: row[c] for c in COLS_PLATAFORMA})

    solo_plataformas = []
    for rut in sorted(solo_plat):
        row = plat_idx.loc[rut]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        solo_plataformas.append({c: row[c] for c in COLS_PLATAFORMA})

    solo_bbdd_rows = []
    for rut in sorted(solo_bbdd):
        row = bbdd_idx.loc[rut]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        solo_bbdd_rows.append({c: row[c] for c in COLS_BBDD})

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_ok = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    fill_plat = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    fill_bbdd = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")

    _escribir_hoja(wb, "Coincidencias", COLS_PLATAFORMA, coincidencias, fill_ok)
    _escribir_hoja(wb, "Solo en Plataformas", COLS_PLATAFORMA, solo_plataformas, fill_plat)
    _escribir_hoja(wb, "Solo BBDD", COLS_BBDD, solo_bbdd_rows, fill_bbdd)

    wb.save(destino)
    return {
        "coincidencias": len(coincidencias),
        "solo_plataformas": len(solo_plataformas),
        "solo_bbdd": len(solo_bbdd_rows),
        "plataforma": len(df_plat),
        "bbdd": len(df_bbdd),
    }


def generar_reporte_bytes(
    ruta_plataforma: Path | str,
    ruta_bbdd: Path | str,
) -> tuple[bytes, str, dict]:
    buf = io.BytesIO()
    stats = procesar_comparacion(ruta_plataforma, ruta_bbdd, buf)
    filename = f"Reporte_Carozzi_Todos_Chile_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf.getvalue(), filename, stats
